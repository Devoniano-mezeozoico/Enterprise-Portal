from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path


try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    import xlrd  # noqa: F401

    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

try:
    import python_calamine  # noqa: F401

    HAS_CALAMINE = True
except ImportError:
    HAS_CALAMINE = False


VERDE = "#2ecc71"
VERMELHO = "#e74c3c"
LARANJA = "#e67e22"
AZUL = "#2980b9"

DOC_RE_NFE = re.compile(
    r"^(\d{2}/\d{2}/\d{2})\s+"
    r"(NFe|NFCo|NF3e|NF)\s+"
    r"(\S+)\s+"
    r"(\d+)\s+"
    r"(\d{2}/\d{2}/\d{2})\s+"
    r"(\S+)\s+"
    r"([A-Z]{2})\s+"
    r"([\d\.,]+)"
    r"(?:\s+(.+))?"
)

DOC_RE_CTE = re.compile(
    r"^(\d{2}/\d{2}/\d{2})\s+"
    r"(CTe|CTE)\s+"
    r"(\S+)\s+"
    r"(\d+)\s+"
    r"(\d{2}/\d{2}/\d{2})\s+"
    r"(\S+)\s+"
    r"([A-Z]{2})\s+"
    r"([\d\.,]+)"
    r"(?:\s+(.+))?"
)

FORN_RE = re.compile(
    r"(?:Fornecedor|Emitente|Transportador):\s+(.+?)\s+CPF/CNPJ:\s+([\d\.\/\-]+)",
    re.IGNORECASE,
)


NFE_MATCH_COLUMNS = (
    ("numero", "Numero NF", 90),
    ("serie", "Serie", 50),
    ("nome", "Emitente", 200),
    ("cnpj", "CNPJ", 130),
    ("uf", "UF", 42),
    ("valor_excel", "Valor Excel", 105),
    ("valor_pdf", "Valor PDF", 105),
    ("dif_valor", "Dif. Valor", 95),
    ("icms_bc_excel", "BC ICMS Excel", 115),
    ("icms_bc_pdf", "BC ICMS PDF", 115),
    ("icms_val_excel", "ICMS Excel", 100),
    ("icms_val_pdf", "ICMS PDF", 100),
    ("ipi_excel", "IPI Excel", 90),
    ("ipi_pdf", "IPI PDF", 90),
    ("tipo_op", "E/S", 42),
    ("situacao", "Situacao", 90),
    ("data_emissao", "Dt Emissao", 90),
    ("data_entrada_pdf", "Dt Entrada PDF", 110),
)

NFE_EXCEL_COLUMNS = (
    ("numero", "Numero NF", 90),
    ("serie", "Serie", 50),
    ("nome", "Emitente", 200),
    ("cnpj", "CNPJ", 130),
    ("uf", "UF", 42),
    ("valor", "Valor", 105),
    ("icms_bc", "BC ICMS", 110),
    ("icms_val", "ICMS", 100),
    ("ipi_val", "IPI", 90),
    ("tipo_op", "E/S", 42),
    ("situacao", "Situacao", 90),
    ("data_emissao", "Dt Emissao", 90),
    ("modelo", "Modelo", 60),
)

NFE_PDF_COLUMNS = (
    ("numero", "Numero NF", 90),
    ("serie", "Serie", 50),
    ("tipo", "Tipo", 60),
    ("fornecedor", "Fornecedor", 200),
    ("cnpj", "CNPJ", 130),
    ("uf", "UF", 42),
    ("valor", "Valor", 105),
    ("icms_bc", "BC ICMS", 110),
    ("icms_val", "ICMS", 100),
    ("ipi_val", "IPI", 90),
    ("data_entrada", "Dt Entrada", 90),
    ("data_doc", "Dt Documento", 90),
)

CTE_MATCH_COLUMNS = (
    ("numero", "Numero CTe", 100),
    ("serie", "Serie", 50),
    ("nome_emit", "Emitente", 200),
    ("cnpj_emit", "CNPJ Emitente", 130),
    ("papel", "Papel Tomador", 110),
    ("valor_excel", "Valor Excel", 105),
    ("valor_pdf", "Valor PDF", 105),
    ("dif_valor", "Dif. Valor", 95),
    ("icms_bc_excel", "BC ICMS Excel", 115),
    ("icms_bc_pdf", "BC ICMS PDF", 115),
    ("icms_val_excel", "ICMS Excel", 100),
    ("icms_val_pdf", "ICMS PDF", 100),
    ("situacao", "Situacao", 90),
    ("data_emissao", "Dt Emissao", 90),
    ("data_entrada_pdf", "Dt Entrada PDF", 110),
)

CTE_EXCEL_COLUMNS = (
    ("numero", "Numero CTe", 100),
    ("serie", "Serie", 50),
    ("nome_emit", "Emitente", 200),
    ("cnpj_emit", "CNPJ Emitente", 130),
    ("papel", "Papel Tomador", 110),
    ("valor", "Valor", 105),
    ("icms_bc", "BC ICMS", 110),
    ("icms_val", "ICMS", 100),
    ("situacao", "Situacao", 90),
    ("data_emissao", "Dt Emissao", 90),
)

CTE_PDF_COLUMNS = (
    ("numero", "Numero CTe", 100),
    ("serie", "Serie", 50),
    ("tipo", "Tipo", 55),
    ("fornecedor", "Fornecedor", 200),
    ("cnpj", "CNPJ", 130),
    ("uf", "UF", 42),
    ("valor", "Valor", 105),
    ("icms_bc", "BC ICMS", 110),
    ("icms_val", "ICMS", 100),
    ("data_entrada", "Dt Entrada", 90),
    ("data_doc", "Dt Documento", 90),
)

TESTE_BATIDOS_COLUMNS = (
    ("numero", "Numero CTe", 130),
    ("val_teste", "Valor Teste", 130),
    ("val_rpt", "Valor RPT", 130),
    ("situacao", "Situacao", 130),
)

TESTE_DIVERGENTES_COLUMNS = (
    ("numero", "Numero CTe", 130),
    ("val_teste", "Valor Teste", 130),
    ("val_rpt", "Valor RPT", 130),
    ("diferenca", "Diferenca", 130),
    ("situacao", "Situacao", 130),
)

TESTE_SIMPLES_COLUMNS = (
    ("numero", "Numero CTe", 150),
    ("valor", "Valor", 150),
    ("situacao", "Situacao", 150),
)


class ConferenciaError(ValueError):
    pass


def dependency_warnings() -> list[str]:
    warnings = []
    if openpyxl is None:
        warnings.append("openpyxl ausente. Instale com: pip install openpyxl")
    if pdfplumber is None:
        warnings.append("pdfplumber ausente. Instale com: pip install pdfplumber")
    if pd is None:
        warnings.append("pandas ausente. Instale com: pip install pandas")
    if not HAS_XLRD:
        warnings.append("xlrd ausente. Necessario para alguns arquivos .xls.")
    return warnings


def require_common_dependencies() -> None:
    missing = []
    if openpyxl is None:
        missing.append("openpyxl")
    if pdfplumber is None:
        missing.append("pdfplumber")
    if missing:
        raise ConferenciaError("Dependencias ausentes: " + ", ".join(missing))


def require_pandas() -> None:
    if pd is None:
        raise ConferenciaError("Dependencia ausente: pandas")


def limpar_cnpj(value):
    if not value:
        return ""
    return re.sub(r"[.\-/\s]", "", str(value)).strip()


def parse_valor_br(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text or text.lower() in ("nan", "none"):
        return 0.0
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except Exception:
        return 0.0


def parse_num_documento(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(round(value)))
    text = str(value).strip()
    if not text or text.lower() in ("nan", "none"):
        return ""
    if "," in text:
        text = text.replace(".", "").replace(",", ".")
    return str(int(round(float(text))))


def money(value) -> str:
    return f"R$ {float(value or 0):,.2f}"


def diff_money(value) -> str:
    if abs(value) < 0.02:
        return "-"
    sign = "+" if value > 0 else "-"
    return f"{sign}R$ {abs(value):,.2f}"


def columns(items):
    return [{"key": key, "label": label, "width": width} for key, label, width in items]


def extrair_excel_nfe(caminho):
    require_common_dependencies()
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    registros = []
    try:
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            try:
                tipo_op = str(row[3]).strip() if row[3] is not None else ""
                situacao = str(row[4]).strip() if row[4] is not None else ""
                data_raw = row[6]
                cnpj = limpar_cnpj(row[8])
                nome = str(row[11]).strip() if row[11] is not None else ""
                uf = str(row[12]).strip() if row[12] is not None else ""
                serie = str(row[19]).strip() if row[19] is not None else ""
                numero = str(row[20]).strip() if row[20] is not None else ""
                valor = float(row[21]) if row[21] is not None else 0.0
                icms_val = float(row[23]) if row[23] is not None else 0.0
                icms_bc = float(row[24]) if row[24] is not None else 0.0
                ipi_val = float(row[43]) if row[43] is not None else 0.0
                modelo = str(row[1]).strip() if row[1] is not None else ""

                if isinstance(data_raw, datetime):
                    data_emissao = data_raw.strftime("%d/%m/%y")
                elif data_raw:
                    data_emissao = str(data_raw)[:10]
                else:
                    data_emissao = ""

                registros.append(
                    {
                        "linha_excel": i + 2,
                        "tipo_op": tipo_op,
                        "situacao": situacao,
                        "data_emissao": data_emissao,
                        "cnpj": cnpj,
                        "numero": numero,
                        "valor": valor,
                        "icms_val": icms_val,
                        "icms_bc": icms_bc,
                        "ipi_val": ipi_val,
                        "nome": nome,
                        "uf": uf,
                        "serie": serie,
                        "modelo": modelo,
                    }
                )
            except Exception:
                continue
    finally:
        wb.close()
    return registros


def extrair_excel_cte(caminho):
    require_common_dependencies()
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = None
    for nome in wb.sheetnames:
        if "cte" in nome.lower() or "encontrada" in nome.lower():
            ws = wb[nome]
            break
    if ws is None:
        ws = wb.active

    registros = []
    try:
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            try:
                situacao = str(row[0]).strip() if row[0] is not None else ""
                serie = str(row[1]).strip() if row[1] is not None else ""
                numero = str(row[2]).strip() if row[2] is not None else ""
                data_raw = row[3]
                nome_emit = str(row[4]).strip() if row[4] is not None else ""
                cnpj_tom = limpar_cnpj(row[5])
                valor = float(row[6]) if row[6] is not None else 0.0
                cnpj_emit = limpar_cnpj(row[12])
                papel = str(row[14]).strip() if row[14] is not None else ""
                icms_bc = float(row[34]) if row[34] is not None else 0.0
                icms_val = float(row[36]) if row[36] is not None else 0.0

                if isinstance(data_raw, datetime):
                    data_emissao = data_raw.strftime("%d/%m/%y")
                elif data_raw:
                    data_emissao = str(data_raw)[:10]
                else:
                    data_emissao = ""

                registros.append(
                    {
                        "linha_excel": i + 2,
                        "situacao": situacao,
                        "serie": serie,
                        "numero": numero,
                        "data_emissao": data_emissao,
                        "nome_emit": nome_emit,
                        "cnpj_emit": cnpj_emit,
                        "cnpj_tom": cnpj_tom,
                        "valor": valor,
                        "papel": papel,
                        "icms_bc": icms_bc,
                        "icms_val": icms_val,
                    }
                )
            except Exception:
                continue
    finally:
        wb.close()
    return registros


def _to_num(text):
    try:
        return float(text.replace(".", "").replace(",", "."))
    except ValueError:
        return None


def parse_icms_ipi_pdf(resto):
    if not resto:
        return 0.0, 0.0, 0.0, 0.0
    text = re.sub(r"\b\d{4,6}\.\d\b", "", resto)
    text = re.sub(r"\b\d\.\d{3}\b", "", text)
    tokens = text.strip().split()
    icms_bc, icms_val = 0.0, 0.0
    ipi_bc, ipi_val = 0.0, 0.0
    section = 0
    i = 0
    while i < len(tokens):
        cod = tokens[i]
        if cod not in ("1", "3"):
            i += 1
            continue
        i += 1
        nums = []
        while i < len(tokens) and tokens[i] not in ("1", "3"):
            num = _to_num(tokens[i])
            if num is not None:
                nums.append(num)
            i += 1
        if cod == "3":
            if section == 0 and nums:
                icms_bc = nums[0]
            elif section == 1 and nums:
                ipi_bc = nums[0]
            section = 1
        else:
            if section == 0:
                if len(nums) >= 3:
                    icms_bc, icms_val = nums[0], nums[2]
                elif len(nums) == 2:
                    icms_bc, icms_val = nums[0], nums[1]
                elif nums:
                    icms_bc = nums[0]
            else:
                if len(nums) >= 2:
                    ipi_bc, ipi_val = nums[0], nums[1]
                elif nums:
                    ipi_bc = nums[0]
            section = 1
    return icms_bc, icms_val, ipi_bc, ipi_val


def extrair_pdf_nfe(caminho, callback_progresso=None):
    require_common_dependencies()
    with pdfplumber.open(caminho) as pdf:
        total_pags = len(pdf.pages) or 1
        linhas = []
        for n, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                linhas.extend(text.split("\n"))
            if callback_progresso:
                callback_progresso(int((n + 1) / total_pags * 50))

    registros = []
    i = 0
    while i < len(linhas):
        line = linhas[i].strip()
        match = DOC_RE_NFE.match(line)
        if match:
            fornec = None
            for j in range(i + 1, min(i + 6, len(linhas))):
                mf = FORN_RE.search(linhas[j])
                if mf:
                    fornec = mf
                    break
            cnpj_clean = limpar_cnpj(fornec.group(2)) if fornec else ""
            icms_bc, icms_val, ipi_bc, ipi_val = parse_icms_ipi_pdf(match.group(9) or "")
            registros.append(
                {
                    "data_entrada": match.group(1),
                    "tipo": match.group(2),
                    "serie": match.group(3),
                    "numero": match.group(4),
                    "data_doc": match.group(5),
                    "uf": match.group(7),
                    "valor": parse_valor_br(match.group(8)),
                    "fornecedor": fornec.group(1) if fornec else "",
                    "cnpj": cnpj_clean,
                    "icms_bc": icms_bc,
                    "icms_val": icms_val,
                    "ipi_bc": ipi_bc,
                    "ipi_val": ipi_val,
                }
            )
        i += 1
    return registros


def parse_icms_pdf_cte(resto):
    if not resto:
        return 0.0, 0.0
    text = re.sub(r"\b\d{4,6}\.\d\b", "", resto)
    text = re.sub(r"\b\d\.\d{3}\b", "", text)
    tokens = text.strip().split()
    icms_bc, icms_val = 0.0, 0.0
    i = 0
    while i < len(tokens):
        cod = tokens[i]
        if cod not in ("1", "3"):
            i += 1
            continue
        i += 1
        nums = []
        while i < len(tokens) and tokens[i] not in ("1", "3"):
            num = _to_num(tokens[i])
            if num is not None:
                nums.append(num)
            i += 1
        if cod == "3" and nums:
            icms_bc = nums[0]
            break
        if cod == "1":
            if len(nums) >= 3:
                icms_bc, icms_val = nums[0], nums[2]
            elif len(nums) == 2:
                icms_bc, icms_val = nums[0], nums[1]
            elif nums:
                icms_bc = nums[0]
            break
    return icms_bc, icms_val


def extrair_pdf_cte(caminho, callback_progresso=None):
    require_common_dependencies()
    with pdfplumber.open(caminho) as pdf:
        total_pags = len(pdf.pages) or 1
        linhas = []
        for n, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                linhas.extend(text.split("\n"))
            if callback_progresso:
                callback_progresso(int((n + 1) / total_pags * 50))

    registros = []
    i = 0
    while i < len(linhas):
        line = linhas[i].strip()
        match = DOC_RE_CTE.match(line)
        if match:
            fornec = None
            for j in range(i + 1, min(i + 6, len(linhas))):
                mf = FORN_RE.search(linhas[j])
                if mf:
                    fornec = mf
                    break
            cnpj_clean = limpar_cnpj(fornec.group(2)) if fornec else ""
            icms_bc, icms_val = parse_icms_pdf_cte(match.group(9) or "")
            registros.append(
                {
                    "data_entrada": match.group(1),
                    "tipo": match.group(2),
                    "serie": match.group(3),
                    "numero": match.group(4),
                    "data_doc": match.group(5),
                    "uf": match.group(7),
                    "valor": parse_valor_br(match.group(8)),
                    "fornecedor": fornec.group(1) if fornec else "",
                    "cnpj": cnpj_clean,
                    "icms_bc": icms_bc,
                    "icms_val": icms_val,
                }
            )
        i += 1
    return registros


def comparar_nfe(excel_rows, pdf_rows, callback_progresso=None):
    pdf_used = set()
    matched = []
    so_excel = []
    total = len(excel_rows) or 1
    for ei, excel_row in enumerate(excel_rows):
        found = False
        for pi, pdf_row in enumerate(pdf_rows):
            if pi in pdf_used:
                continue
            cnpj_ok = (
                excel_row["cnpj"][:8] == pdf_row["cnpj"][:8]
                if (excel_row["cnpj"] and pdf_row["cnpj"])
                else False
            )
            num_ok = excel_row["numero"] == pdf_row["numero"]
            valor_ok = abs(excel_row["valor"] - pdf_row["valor"]) < 0.02
            if cnpj_ok and num_ok and valor_ok:
                matched.append({"excel": excel_row, "pdf": pdf_row})
                pdf_used.add(pi)
                found = True
                break
        if not found:
            so_excel.append(excel_row)
        if callback_progresso:
            callback_progresso(50 + int((ei + 1) / total * 50))
    so_pdf = [pdf_row for pi, pdf_row in enumerate(pdf_rows) if pi not in pdf_used]
    return matched, so_excel, so_pdf


def comparar_cte(excel_rows, pdf_rows, callback_progresso=None):
    pdf_used = set()
    matched = []
    so_excel = []
    total = len(excel_rows) or 1
    for ei, excel_row in enumerate(excel_rows):
        found = False
        for pi, pdf_row in enumerate(pdf_rows):
            if pi in pdf_used:
                continue
            cnpj_e = excel_row["cnpj_emit"] or excel_row["cnpj_tom"]
            cnpj_p = pdf_row["cnpj"]
            cnpj_ok = (cnpj_e[:8] == cnpj_p[:8]) if (cnpj_e and cnpj_p) else False
            num_ok = excel_row["numero"] == pdf_row["numero"]
            valor_ok = abs(excel_row["valor"] - pdf_row["valor"]) < 0.02
            if cnpj_ok and num_ok and valor_ok:
                matched.append({"excel": excel_row, "pdf": pdf_row})
                pdf_used.add(pi)
                found = True
                break
        if not found:
            so_excel.append(excel_row)
        if callback_progresso:
            callback_progresso(50 + int((ei + 1) / total * 50))
    so_pdf = [pdf_row for pi, pdf_row in enumerate(pdf_rows) if pi not in pdf_used]
    return matched, so_excel, so_pdf


def _ler_arquivo_generico(caminho):
    require_pandas()
    ext = os.path.splitext(caminho)[1].lower()

    if ext == ".csv":
        for enc in ("utf-8-sig", "latin1", "cp1252"):
            try:
                return pd.read_csv(caminho, encoding=enc, sep=None, engine="python")
            except Exception:
                continue
        raise ConferenciaError(f"Nao foi possivel ler o arquivo CSV: {caminho}")

    if ext in (".xlsx", ".xlsm"):
        return pd.read_excel(caminho, engine="openpyxl")

    if ext == ".xls":
        if HAS_XLRD:
            try:
                return pd.read_excel(caminho, engine="xlrd")
            except Exception:
                pass
        if HAS_CALAMINE:
            try:
                return pd.read_excel(caminho, engine="calamine")
            except Exception:
                pass
        try:
            dfs = pd.read_html(caminho, decimal=",", thousands=".")
        except Exception:
            dfs = None
        if not dfs:
            for enc in ("latin1", "cp1252", "utf-8"):
                try:
                    dfs = pd.read_html(caminho, encoding=enc, decimal=",", thousands=".")
                    if dfs:
                        break
                except Exception:
                    continue
        if not dfs:
            raise ConferenciaError(
                "Nao foi possivel ler o arquivo XLS. Instale xlrd ou confira se o arquivo nao esta corrompido."
            )
        df = dfs[0]
        if df.iloc[0].apply(lambda x: isinstance(x, str)).all():
            df.columns = df.iloc[0]
            df = df.drop(0).reset_index(drop=True)
        return df

    return pd.read_excel(caminho, engine="openpyxl")


def extrair_teste(caminho):
    df = _ler_arquivo_generico(caminho)
    df.columns = [str(c).strip() for c in df.columns]
    col_num = next(
        (
            c
            for c in df.columns
            if "nr" in c.lower()
            or "numero" in c.lower()
            or "número" in c.lower()
            or "documento" in c.lower()
        ),
        None,
    )
    col_val = next(
        (
            c
            for c in df.columns
            if "valor" in c.lower() or "contabil" in c.lower() or "contábil" in c.lower()
        ),
        None,
    )
    if col_num is None or col_val is None:
        raise ConferenciaError(f"Colunas esperadas nao encontradas no Teste. Colunas: {list(df.columns)}")

    registros = []
    for _, row in df.iterrows():
        try:
            num = parse_num_documento(row[col_num])
            if not num:
                continue
            registros.append({"numero": num, "valor": parse_valor_br(row[col_val])})
        except Exception:
            continue
    return registros


def extrair_rpt(caminho):
    df = _ler_arquivo_generico(caminho)
    df.columns = [str(c).strip() for c in df.columns]

    if df.columns.tolist() and all(isinstance(c, (int, float)) for c in df.columns[:3]):
        real_headers = [str(h).strip() for h in df.iloc[0].tolist()]
        df.columns = real_headers
        df = df.drop(0).reset_index(drop=True)

    col_num = None
    col_val = None
    col_sit = None
    idx_sit, idx_num, idx_val = 5, 24, 42
    if len(df.columns) > max(idx_sit, idx_num, idx_val):
        col_sit = df.columns[idx_sit]
        col_num = df.columns[idx_num]
        col_val = df.columns[idx_val]

    cols = list(df.columns)
    if col_num is None or col_val is None:
        col_num = next(
            (
                c
                for c in cols
                if "NUMERO_CTE" in c.upper().replace(" ", "_")
                or "NÚMERO_CTE" in c.upper().replace(" ", "_")
                or ("N" in c.upper() and "CTE" in c.upper())
            ),
            None,
        )
        col_val = next(
            (
                c
                for c in cols
                if "VALOR_TOTAL_PREST" in c.upper().replace(" ", "_")
                or "VALOR_TOTAL" in c.upper().replace(" ", "_")
            ),
            None,
        )
    if col_sit is None:
        col_sit = next(
            (
                c
                for c in cols
                if "SITUACAO" in c.upper().replace(" ", "_")
                or "SITUAÇÃO" in c.upper().replace(" ", "_")
            ),
            None,
        )

    if col_num is None or col_val is None:
        raise ConferenciaError(f"Colunas esperadas nao encontradas no RPT. Colunas: {list(df.columns[:30])}")

    registros = []
    for _, row in df.iterrows():
        try:
            num = parse_num_documento(row[col_num])
            if not num:
                continue
            val = parse_valor_br(row[col_val])
            sit = str(row[col_sit]).strip() if (col_sit is not None and row[col_sit] is not None) else ""
            if sit.lower() in ("nan", "none"):
                sit = ""
            registros.append({"numero": num, "valor": val, "situacao": sit})
        except Exception:
            continue
    return registros


def comparar_teste_rpt(teste_rows, rpt_rows):
    rpt_map = {}
    for row in rpt_rows:
        rpt_map.setdefault(row["numero"], []).append(row)

    batidos = []
    divergentes = []
    so_teste = []
    teste_usados = set()

    for teste in teste_rows:
        num = teste["numero"]
        if num in rpt_map:
            rpt_regs = rpt_map[num]
            matched_reg = None
            for rpt in rpt_regs:
                if abs(teste["valor"] - rpt["valor"]) < 0.02:
                    matched_reg = rpt
                    break
            if matched_reg is not None:
                batidos.append(
                    {
                        "numero": num,
                        "val_teste": teste["valor"],
                        "val_rpt": matched_reg["valor"],
                        "situacao": matched_reg.get("situacao", ""),
                    }
                )
            else:
                divergentes.append(
                    {
                        "numero": num,
                        "val_teste": teste["valor"],
                        "val_rpt": rpt_regs[0]["valor"],
                        "situacao": rpt_regs[0].get("situacao", ""),
                    }
                )
            teste_usados.add(num)
        else:
            so_teste.append({"numero": num, "valor": teste["valor"], "situacao": ""})

    so_rpt = [
        {"numero": row["numero"], "valor": row["valor"], "situacao": row.get("situacao", "")}
        for row in rpt_rows
        if row["numero"] not in teste_usados
    ]
    return batidos, so_teste, so_rpt, divergentes


def _row_class_for_diff(diff):
    if abs(diff) < 0.02:
        return ""
    return "dif-pos" if diff > 0 else "dif-neg"


def _tab(tab_id, title, color, columns_def, rows):
    return {
        "id": tab_id,
        "title": title,
        "color": color,
        "columns": columns(columns_def),
        "rows": rows,
        "count": len(rows),
    }


def build_document_tabs(matched, so_excel, so_pdf, modo):
    if modo == "nfe":
        match_cols, excel_cols, pdf_cols = NFE_MATCH_COLUMNS, NFE_EXCEL_COLUMNS, NFE_PDF_COLUMNS
    else:
        match_cols, excel_cols, pdf_cols = CTE_MATCH_COLUMNS, CTE_EXCEL_COLUMNS, CTE_PDF_COLUMNS

    match_rows = []
    for item in matched:
        excel = item["excel"]
        pdf = item["pdf"]
        diff = pdf["valor"] - excel["valor"]
        if modo == "nfe":
            row = {
                "row_class": _row_class_for_diff(diff),
                "numero": excel["numero"],
                "serie": excel["serie"],
                "nome": excel["nome"],
                "cnpj": excel["cnpj"],
                "uf": excel["uf"] or pdf["uf"],
                "valor_excel": money(excel["valor"]),
                "valor_pdf": money(pdf["valor"]),
                "dif_valor": diff_money(diff),
                "icms_bc_excel": money(excel["icms_bc"]),
                "icms_bc_pdf": money(pdf["icms_bc"]),
                "icms_val_excel": money(excel["icms_val"]),
                "icms_val_pdf": money(pdf["icms_val"]),
                "ipi_excel": money(excel["ipi_val"]),
                "ipi_pdf": money(pdf["ipi_val"]),
                "tipo_op": excel["tipo_op"],
                "situacao": excel["situacao"],
                "data_emissao": excel["data_emissao"],
                "data_entrada_pdf": pdf["data_entrada"],
            }
        else:
            row = {
                "row_class": _row_class_for_diff(diff),
                "numero": excel["numero"],
                "serie": excel["serie"],
                "nome_emit": excel["nome_emit"],
                "cnpj_emit": excel["cnpj_emit"],
                "papel": excel["papel"],
                "valor_excel": money(excel["valor"]),
                "valor_pdf": money(pdf["valor"]),
                "dif_valor": diff_money(diff),
                "icms_bc_excel": money(excel["icms_bc"]),
                "icms_bc_pdf": money(pdf["icms_bc"]),
                "icms_val_excel": money(excel["icms_val"]),
                "icms_val_pdf": money(pdf["icms_val"]),
                "situacao": excel["situacao"],
                "data_emissao": excel["data_emissao"],
                "data_entrada_pdf": pdf["data_entrada"],
            }
        match_rows.append(row)

    excel_rows = []
    for excel in so_excel:
        row = {"row_class": "cancelado" if str(excel.get("situacao", "")).upper() == "CANCELADO" else ""}
        if modo == "nfe":
            row.update(
                {
                    "numero": excel["numero"],
                    "serie": excel["serie"],
                    "nome": excel["nome"],
                    "cnpj": excel["cnpj"],
                    "uf": excel["uf"],
                    "valor": money(excel["valor"]),
                    "icms_bc": money(excel["icms_bc"]),
                    "icms_val": money(excel["icms_val"]),
                    "ipi_val": money(excel["ipi_val"]),
                    "tipo_op": excel["tipo_op"],
                    "situacao": excel["situacao"],
                    "data_emissao": excel["data_emissao"],
                    "modelo": excel["modelo"],
                }
            )
        else:
            row.update(
                {
                    "numero": excel["numero"],
                    "serie": excel["serie"],
                    "nome_emit": excel["nome_emit"],
                    "cnpj_emit": excel["cnpj_emit"],
                    "papel": excel["papel"],
                    "valor": money(excel["valor"]),
                    "icms_bc": money(excel["icms_bc"]),
                    "icms_val": money(excel["icms_val"]),
                    "situacao": excel["situacao"],
                    "data_emissao": excel["data_emissao"],
                }
            )
        excel_rows.append(row)

    pdf_rows = []
    for pdf in so_pdf:
        row = {"row_class": ""}
        if modo == "nfe":
            row.update(
                {
                    "numero": pdf["numero"],
                    "serie": pdf["serie"],
                    "tipo": pdf["tipo"],
                    "fornecedor": pdf["fornecedor"],
                    "cnpj": pdf["cnpj"],
                    "uf": pdf["uf"],
                    "valor": money(pdf["valor"]),
                    "icms_bc": money(pdf["icms_bc"]),
                    "icms_val": money(pdf["icms_val"]),
                    "ipi_val": money(pdf["ipi_val"]),
                    "data_entrada": pdf["data_entrada"],
                    "data_doc": pdf["data_doc"],
                }
            )
        else:
            row.update(
                {
                    "numero": pdf["numero"],
                    "serie": pdf["serie"],
                    "tipo": pdf["tipo"],
                    "fornecedor": pdf["fornecedor"],
                    "cnpj": pdf["cnpj"],
                    "uf": pdf["uf"],
                    "valor": money(pdf["valor"]),
                    "icms_bc": money(pdf["icms_bc"]),
                    "icms_val": money(pdf["icms_val"]),
                    "data_entrada": pdf["data_entrada"],
                    "data_doc": pdf["data_doc"],
                }
            )
        pdf_rows.append(row)

    valor_conferido = sum(item["excel"]["valor"] for item in matched)
    valor_sem_pdf = sum(item["valor"] for item in so_excel)
    title = "Conferencia de Entradas Fiscais - Modelo 55 (NF-e)" if modo == "nfe" else "Conferencia de CTe's Fiscais"

    return {
        "kind": "documento",
        "mode": modo,
        "title": title,
        "tabs": [
            _tab("match", "Encontrados nos dois", VERDE, match_cols, match_rows),
            _tab("excel", "So no Excel (SAT)", VERMELHO, excel_cols, excel_rows),
            _tab("pdf", "So no PDF", LARANJA, pdf_cols, pdf_rows),
        ],
        "summary": {
            "matched": len(matched),
            "so_excel": len(so_excel),
            "so_pdf": len(so_pdf),
            "valor_conferido": money(valor_conferido),
            "valor_sem_pdf": money(valor_sem_pdf),
        },
    }


def build_teste_tabs(batidos, so_teste, so_rpt, divergentes):
    batidos_rows = [
        {
            "row_class": "",
            "numero": row["numero"],
            "val_teste": money(row["val_teste"]),
            "val_rpt": money(row["val_rpt"]),
            "situacao": row.get("situacao", ""),
        }
        for row in batidos
    ]
    divergentes_rows = [
        {
            "row_class": "dif-neg",
            "numero": row["numero"],
            "val_teste": money(row["val_teste"]),
            "val_rpt": money(row["val_rpt"]),
            "diferenca": money(round(row["val_teste"] - row["val_rpt"], 2)),
            "situacao": row.get("situacao", ""),
        }
        for row in divergentes
    ]
    so_teste_rows = [
        {"row_class": "", "numero": row["numero"], "valor": money(row["valor"]), "situacao": row.get("situacao", "")}
        for row in so_teste
    ]
    so_rpt_rows = [
        {"row_class": "", "numero": row["numero"], "valor": money(row["valor"]), "situacao": row.get("situacao", "")}
        for row in so_rpt
    ]

    return {
        "kind": "teste",
        "mode": "teste",
        "title": "Comparacao Teste x RPT",
        "tabs": [
            _tab("batidos", "Valores que batem", VERDE, TESTE_BATIDOS_COLUMNS, batidos_rows),
            _tab("divergentes", "Valores divergentes", VERMELHO, TESTE_DIVERGENTES_COLUMNS, divergentes_rows),
            _tab("so_teste", "So no Teste", LARANJA, TESTE_SIMPLES_COLUMNS, so_teste_rows),
            _tab("so_rpt", "So no RPT", AZUL, TESTE_SIMPLES_COLUMNS, so_rpt_rows),
        ],
        "summary": {
            "batidos": len(batidos),
            "divergentes": len(divergentes),
            "so_teste": len(so_teste),
            "so_rpt": len(so_rpt),
        },
    }


def processar_documento(modo, excel_path, pdf_path):
    if modo not in ("nfe", "cte"):
        raise ConferenciaError("Modo invalido.")
    if not Path(excel_path).is_file():
        raise ConferenciaError("Selecione o arquivo Excel SAT.")
    if not Path(pdf_path).is_file():
        raise ConferenciaError("Selecione o arquivo PDF.")

    if modo == "nfe":
        excel_rows = extrair_excel_nfe(excel_path)
        pdf_rows = extrair_pdf_nfe(pdf_path)
        matched, so_excel, so_pdf = comparar_nfe(excel_rows, pdf_rows)
    else:
        excel_rows = extrair_excel_cte(excel_path)
        pdf_rows = extrair_pdf_cte(pdf_path)
        matched, so_excel, so_pdf = comparar_cte(excel_rows, pdf_rows)

    display = build_document_tabs(matched, so_excel, so_pdf, modo)
    display["raw"] = {"matched": matched, "so_excel": so_excel, "so_pdf": so_pdf, "modo": modo}
    return display


def processar_teste(teste_path, rpt_path):
    if not Path(teste_path).is_file():
        raise ConferenciaError("Selecione o arquivo Teste.")
    if not Path(rpt_path).is_file():
        raise ConferenciaError("Selecione o arquivo RPT.")
    teste_rows = extrair_teste(teste_path)
    rpt_rows = extrair_rpt(rpt_path)
    batidos, so_teste, so_rpt, divergentes = comparar_teste_rpt(teste_rows, rpt_rows)
    display = build_teste_tabs(batidos, so_teste, so_rpt, divergentes)
    display["raw"] = {
        "batidos": batidos,
        "so_teste": so_teste,
        "so_rpt": so_rpt,
        "divergentes": divergentes,
    }
    return display


def exportar_documento(matched, so_excel, so_pdf, modo, output_path):
    require_common_dependencies()
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    if modo == "nfe":
        ws1 = wb.active
        ws1.title = "Conferidos"
        ws1.append(
            [
                "Numero NF",
                "Serie",
                "Emitente",
                "CNPJ",
                "UF",
                "Valor Excel",
                "Valor PDF",
                "Dif. Valor (PDF-Excel)",
                "BC ICMS Excel",
                "BC ICMS PDF",
                "ICMS Excel",
                "ICMS PDF",
                "IPI Excel",
                "IPI PDF",
                "E/S",
                "Situacao",
                "Dt Emissao",
                "Dt Entrada PDF",
            ]
        )
        for item in matched:
            excel, pdf = item["excel"], item["pdf"]
            ws1.append(
                [
                    excel["numero"],
                    excel["serie"],
                    excel["nome"],
                    excel["cnpj"],
                    excel["uf"] or pdf["uf"],
                    excel["valor"],
                    pdf["valor"],
                    round(pdf["valor"] - excel["valor"], 2),
                    excel["icms_bc"],
                    pdf["icms_bc"],
                    excel["icms_val"],
                    pdf["icms_val"],
                    excel["ipi_val"],
                    pdf["ipi_val"],
                    excel["tipo_op"],
                    excel["situacao"],
                    excel["data_emissao"],
                    pdf["data_entrada"],
                ]
            )
        ws2 = wb.create_sheet("So no Excel")
        ws2.append(["Numero NF", "Serie", "Emitente", "CNPJ", "UF", "Valor", "BC ICMS", "ICMS", "IPI", "E/S", "Situacao", "Dt Emissao", "Modelo"])
        for excel in so_excel:
            ws2.append([excel["numero"], excel["serie"], excel["nome"], excel["cnpj"], excel["uf"], excel["valor"], excel["icms_bc"], excel["icms_val"], excel["ipi_val"], excel["tipo_op"], excel["situacao"], excel["data_emissao"], excel["modelo"]])
        ws3 = wb.create_sheet("So no PDF")
        ws3.append(["Numero NF", "Serie", "Tipo", "Fornecedor", "CNPJ", "UF", "Valor", "BC ICMS", "ICMS", "IPI", "Dt Entrada", "Dt Documento"])
        for pdf in so_pdf:
            ws3.append([pdf["numero"], pdf["serie"], pdf["tipo"], pdf["fornecedor"], pdf["cnpj"], pdf["uf"], pdf["valor"], pdf["icms_bc"], pdf["icms_val"], pdf["ipi_val"], pdf["data_entrada"], pdf["data_doc"]])
    else:
        ws1 = wb.active
        ws1.title = "Conferidos"
        ws1.append(["Numero CTe", "Serie", "Emitente", "CNPJ Emitente", "Papel Tomador", "Valor Excel", "Valor PDF", "Dif. Valor (PDF-Excel)", "BC ICMS Excel", "BC ICMS PDF", "ICMS Excel", "ICMS PDF", "Situacao", "Dt Emissao", "Dt Entrada PDF"])
        for item in matched:
            excel, pdf = item["excel"], item["pdf"]
            ws1.append([excel["numero"], excel["serie"], excel["nome_emit"], excel["cnpj_emit"], excel["papel"], excel["valor"], pdf["valor"], round(pdf["valor"] - excel["valor"], 2), excel["icms_bc"], pdf["icms_bc"], excel["icms_val"], pdf["icms_val"], excel["situacao"], excel["data_emissao"], pdf["data_entrada"]])
        ws2 = wb.create_sheet("So no Excel")
        ws2.append(["Numero CTe", "Serie", "Emitente", "CNPJ Emitente", "Papel Tomador", "Valor", "BC ICMS", "ICMS", "Situacao", "Dt Emissao"])
        for excel in so_excel:
            ws2.append([excel["numero"], excel["serie"], excel["nome_emit"], excel["cnpj_emit"], excel["papel"], excel["valor"], excel["icms_bc"], excel["icms_val"], excel["situacao"], excel["data_emissao"]])
        ws3 = wb.create_sheet("So no PDF")
        ws3.append(["Numero CTe", "Serie", "Tipo", "Fornecedor", "CNPJ", "UF", "Valor", "BC ICMS", "ICMS", "Dt Entrada", "Dt Documento"])
        for pdf in so_pdf:
            ws3.append([pdf["numero"], pdf["serie"], pdf["tipo"], pdf["fornecedor"], pdf["cnpj"], pdf["uf"], pdf["valor"], pdf["icms_bc"], pdf["icms_val"], pdf["data_entrada"], pdf["data_doc"]])

    _style_workbook(wb, [ws1, ws2, ws3], ["1B5E20", "B71C1C", "E65100"])
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def exportar_teste(batidos, so_teste, so_rpt, divergentes, output_path):
    require_common_dependencies()
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()

    def escrever_aba(ws, linhas, cabecalho):
        ws.append(cabecalho)
        for row in linhas:
            ws.append(row)

    ws1 = wb.active
    ws1.title = "Valores que batem"
    escrever_aba(ws1, [[d["numero"], d["val_teste"], d["val_rpt"], d.get("situacao", "")] for d in batidos], ["Numero CTe", "Valor Teste", "Valor RPT", "Situacao"])

    ws2 = wb.create_sheet("Valores divergentes")
    escrever_aba(ws2, [[d["numero"], d["val_teste"], d["val_rpt"], round(d["val_teste"] - d["val_rpt"], 2), d.get("situacao", "")] for d in divergentes], ["Numero CTe", "Valor Teste", "Valor RPT", "Diferenca", "Situacao"])

    ws3 = wb.create_sheet("So no Teste")
    escrever_aba(ws3, [[d["numero"], d["valor"], d.get("situacao", "")] for d in so_teste], ["Numero CTe", "Valor Teste", "Situacao"])

    ws4 = wb.create_sheet("So no RPT")
    escrever_aba(ws4, [[d["numero"], d["valor"], d.get("situacao", "")] for d in so_rpt], ["Numero CTe", "Valor RPT", "Situacao"])

    _style_workbook(wb, [ws1, ws2, ws3, ws4], ["1B5E20", "B71C1C", "E65100", "1565C0"], max_width=40)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _style_workbook(wb, sheets, colors, max_width=52):
    from openpyxl.styles import Alignment, Font, PatternFill

    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")
    for ws, color in zip(sheets, colors):
        fill = PatternFill("solid", fgColor=color)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = header_font
            cell.alignment = center
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, max_width)
