import re
import os
import sys
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime

try:
    import openpyxl
except ImportError:
    messagebox.showerror("Dependência ausente", "Instale openpyxl:\n  pip install openpyxl")
    sys.exit(1)

try:
    import pdfplumber
except ImportError:
    messagebox.showerror("Dependência ausente", "Instale pdfplumber:\n  pip install pdfplumber")
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    messagebox.showerror("Dependência ausente", "Instale pandas:\n  pip install pandas")
    sys.exit(1)

# Engines opcionais para leitura de XLS binário legado
try:
    import xlrd  # noqa: F401
    _TEM_XLRD = True
except ImportError:
    _TEM_XLRD = False

try:
    import python_calamine  # noqa: F401
    _TEM_CALAMINE = True
except ImportError:
    _TEM_CALAMINE = False


BASE_DIR  = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
POPPLER   = os.path.join(BASE_DIR, "ZIPS", "poppler-26.02.0", "Library", "bin")
TESSERACT = os.path.join(BASE_DIR, "Tesseract-OCR", "tesseract.exe")

if os.path.isdir(POPPLER):
    os.environ["PATH"] = POPPLER + os.pathsep + os.environ.get("PATH", "")
if os.path.isfile(TESSERACT):
    os.environ["TESSERACT_CMD"] = TESSERACT
    # Adiciona o diretório do Tesseract ao PATH para que o executável seja encontrado
    tesseract_dir = os.path.dirname(TESSERACT)
    os.environ["PATH"] = tesseract_dir + os.pathsep + os.environ.get("PATH", "")
    # Configura pytesseract se estiver disponível
    try:
        import pytesseract
        pytesseract.pytesseract.tesseract_cmd = TESSERACT
    except ImportError:
        pass


def limpar_cnpj(v):
    if not v:
        return ""
    return re.sub(r"[.\-/\s]", "", str(v)).strip()


def parse_valor_br(s):
    """Converte para float. Aceita números nativos (int/float) e strings
    no formato brasileiro ('1.234,56') ou já em formato padrão ('1234.56')."""
    if s is None:
        return 0.0
    if isinstance(s, (int, float)):
        return float(s)
    txt = str(s).strip()
    if not txt or txt.lower() in ("nan", "none"):
        return 0.0
    # Se tem vírgula, assume formato BR: ponto = milhar, vírgula = decimal
    if "," in txt:
        txt = txt.replace(".", "").replace(",", ".")
    # Caso contrário, já está em formato padrão (ponto decimal) — não altera
    try:
        return float(txt)
    except Exception:
        return 0.0


def parse_num_documento(v):
    """Converte um número de documento/CTe para string de inteiro,
    aceitando números nativos (int/float) ou strings em formato BR."""
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return str(int(round(v)))
    txt = str(v).strip()
    if not txt or txt.lower() in ("nan", "none"):
        return ""
    if "," in txt:
        txt = txt.replace(".", "").replace(",", ".")
    return str(int(round(float(txt))))


# ---------------------------------------------------------------------------
# NF-e Modelo 55 — Excel (ENTRADAS SAT)
# Colunas relevantes (0-based):
#   [3]  TipoDeOperacaoEntradaOuSaida
#   [4]  Situacao
#   [6]  DataEmissao
#   [8]  CnpjDoEmitente
#   [11] NomeEmitente
#   [12] UfEmitente
#   [19] SerieDocumento
#   [20] NumeroDocumento
#   [21] ValorTotalNota  ← amarela
#   [23] ICMS / ValorTotalICMS
#   [24] BC ICMS / ValorBaseCalculoICMS
#   [43] IPI / ValorIPI
# ---------------------------------------------------------------------------
def extrair_excel_nfe(caminho):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    registros = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        try:
            tipo_op  = str(row[3]).strip()  if row[3]  is not None else ""
            situacao = str(row[4]).strip()  if row[4]  is not None else ""
            data_raw = row[6]
            cnpj     = limpar_cnpj(row[8])
            nome     = str(row[11]).strip() if row[11] is not None else ""
            uf       = str(row[12]).strip() if row[12] is not None else ""
            serie    = str(row[19]).strip() if row[19] is not None else ""
            numero   = str(row[20]).strip() if row[20] is not None else ""
            valor    = float(row[21])       if row[21] is not None else 0.0
            icms_val = float(row[23])       if row[23] is not None else 0.0
            icms_bc  = float(row[24])       if row[24] is not None else 0.0
            ipi_val  = float(row[43])       if row[43] is not None else 0.0
            modelo   = str(row[1]).strip()  if row[1]  is not None else ""

            if isinstance(data_raw, datetime):
                data_emissao = data_raw.strftime("%d/%m/%y")
            elif data_raw:
                data_emissao = str(data_raw)[:10]
            else:
                data_emissao = ""

            registros.append({
                "linha_excel":  i + 2,
                "tipo_op":      tipo_op,
                "situacao":     situacao,
                "data_emissao": data_emissao,
                "cnpj":         cnpj,
                "numero":       numero,
                "valor":        valor,
                "icms_val":     icms_val,
                "icms_bc":      icms_bc,
                "ipi_val":      ipi_val,
                "nome":         nome,
                "uf":           uf,
                "serie":        serie,
                "modelo":       modelo,
            })
        except Exception:
            continue
    wb.close()
    return registros


# ---------------------------------------------------------------------------
# CTe — Excel (SAT CTE TOMADOS)
# Colunas amarelas (0-based):
#   [0]  Situação
#   [2]  Número CTe
#   [3]  Data emissão
#   [6]  Valor total prestação
#   [12] CNPJ emitente
#   [14] Papel tomador
#   [34] Valor BC ICMS
#   [36] Valor ICMS
# Extras úteis:
#   [1]  Série
#   [4]  Nome emitente
#   [5]  CNPJ/CPF tomador
# ---------------------------------------------------------------------------
def extrair_excel_cte(caminho):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = None
    for nome in wb.sheetnames:
        if "cte" in nome.lower() or "encontrada" in nome.lower():
            ws = wb[nome]
            break
    if ws is None:
        ws = wb.active

    registros = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        try:
            situacao  = str(row[0]).strip()  if row[0]  is not None else ""
            serie     = str(row[1]).strip()  if row[1]  is not None else ""
            numero    = str(row[2]).strip()  if row[2]  is not None else ""
            data_raw  = row[3]
            nome_emit = str(row[4]).strip()  if row[4]  is not None else ""
            cnpj_tom  = limpar_cnpj(row[5])
            valor     = float(row[6])        if row[6]  is not None else 0.0
            cnpj_emit = limpar_cnpj(row[12])
            papel     = str(row[14]).strip() if row[14] is not None else ""
            icms_bc   = float(row[34])       if row[34] is not None else 0.0
            icms_val  = float(row[36])       if row[36] is not None else 0.0

            if isinstance(data_raw, datetime):
                data_emissao = data_raw.strftime("%d/%m/%y")
            elif data_raw:
                data_emissao = str(data_raw)[:10]
            else:
                data_emissao = ""

            registros.append({
                "linha_excel":  i + 2,
                "situacao":     situacao,
                "serie":        serie,
                "numero":       numero,
                "data_emissao": data_emissao,
                "nome_emit":    nome_emit,
                "cnpj_emit":    cnpj_emit,
                "cnpj_tom":     cnpj_tom,
                "valor":        valor,
                "papel":        papel,
                "icms_bc":      icms_bc,
                "icms_val":     icms_val,
            })
        except Exception:
            continue
    wb.close()
    return registros


# ---------------------------------------------------------------------------
# PDF parsing — NF-e (Modelo 55)
# ---------------------------------------------------------------------------
DOC_RE_NFE = re.compile(
    r"^(\d{2}/\d{2}/\d{2})\s+"
    r"(NFe|NFCo|NF3e|NF)\s+"
    r"(\S+)\s+"
    r"(\d+)\s+"
    r"(\d{2}/\d{2}/\d{2})\s+"
    r"(\S+)\s+"
    r"([A-Z]{2})\s+"
    r"([\d\.,]+)"
    r"(?:\s+(.+))?"
)
FORN_RE = re.compile(
    r"(?:Fornecedor|Emitente|Transportador):\s+(.+?)\s+CPF/CNPJ:\s+([\d\.\/\-]+)",
    re.IGNORECASE
)


def _to_num(t):
    try:
        return float(t.replace(".", "").replace(",", "."))
    except ValueError:
        return None


def parse_icms_ipi_pdf(resto):
    if not resto:
        return 0.0, 0.0, 0.0, 0.0
    texto = re.sub(r"\b\d{4,6}\.\d\b", "", resto)
    texto = re.sub(r"\b\d\.\d{3}\b", "", texto)
    tokens = texto.strip().split()
    icms_bc, icms_val = 0.0, 0.0
    ipi_bc,  ipi_val  = 0.0, 0.0
    section = 0
    i = 0
    while i < len(tokens):
        cod = tokens[i]
        if cod not in ("1", "3"):
            i += 1
            continue
        i += 1
        nums = []
        while i < len(tokens) and tokens[i] not in ("1", "3"):
            n = _to_num(tokens[i])
            if n is not None:
                nums.append(n)
            i += 1
        if cod == "3":
            if section == 0 and nums:
                icms_bc = nums[0]
            elif section == 1 and nums:
                ipi_bc = nums[0]
            section = 1
        else:
            if section == 0:
                if len(nums) >= 3:
                    icms_bc, icms_val = nums[0], nums[2]
                elif len(nums) == 2:
                    icms_bc, icms_val = nums[0], nums[1]
                elif nums:
                    icms_bc = nums[0]
            else:
                if len(nums) >= 2:
                    ipi_bc, ipi_val = nums[0], nums[1]
                elif nums:
                    ipi_bc = nums[0]
            section = 1
    return icms_bc, icms_val, ipi_bc, ipi_val


def extrair_pdf_nfe(caminho, callback_progresso=None):
    with pdfplumber.open(caminho) as pdf:
        total_pags = len(pdf.pages)
        linhas = []
        for n, page in enumerate(pdf.pages):
            texto = page.extract_text()
            if texto:
                linhas.extend(texto.split("\n"))
            if callback_progresso:
                callback_progresso(int((n + 1) / total_pags * 50))

    registros = []
    i = 0
    while i < len(linhas):
        linha = linhas[i].strip()
        m = DOC_RE_NFE.match(linha)
        if m:
            fornec = None
            for j in range(i + 1, min(i + 6, len(linhas))):
                mf = FORN_RE.search(linhas[j])
                if mf:
                    fornec = mf
                    break
            cnpj_clean = limpar_cnpj(fornec.group(2)) if fornec else ""
            icms_bc, icms_val, ipi_bc, ipi_val = parse_icms_ipi_pdf(m.group(9) or "")
            registros.append({
                "data_entrada": m.group(1),
                "tipo":         m.group(2),
                "serie":        m.group(3),
                "numero":       m.group(4),
                "data_doc":     m.group(5),
                "uf":           m.group(7),
                "valor":        parse_valor_br(m.group(8)),
                "fornecedor":   fornec.group(1) if fornec else "",
                "cnpj":         cnpj_clean,
                "icms_bc":      icms_bc,
                "icms_val":     icms_val,
                "ipi_bc":       ipi_bc,
                "ipi_val":      ipi_val,
            })
        i += 1
    return registros


# ---------------------------------------------------------------------------
# PDF parsing — CTe
# ---------------------------------------------------------------------------
DOC_RE_CTE = re.compile(
    r"^(\d{2}/\d{2}/\d{2})\s+"
    r"(CTe|CTE)\s+"
    r"(\S+)\s+"
    r"(\d+)\s+"
    r"(\d{2}/\d{2}/\d{2})\s+"
    r"(\S+)\s+"
    r"([A-Z]{2})\s+"
    r"([\d\.,]+)"
    r"(?:\s+(.+))?"
)


def parse_icms_pdf_cte(resto):
    if not resto:
        return 0.0, 0.0
    texto = re.sub(r"\b\d{4,6}\.\d\b", "", resto)
    texto = re.sub(r"\b\d\.\d{3}\b", "", texto)
    tokens = texto.strip().split()
    icms_bc, icms_val = 0.0, 0.0
    i = 0
    while i < len(tokens):
        cod = tokens[i]
        if cod not in ("1", "3"):
            i += 1
            continue
        i += 1
        nums = []
        while i < len(tokens) and tokens[i] not in ("1", "3"):
            n = _to_num(tokens[i])
            if n is not None:
                nums.append(n)
            i += 1
        if cod == "3" and nums:
            icms_bc = nums[0]
            break
        elif cod == "1":
            if len(nums) >= 3:
                icms_bc  = nums[0]
                icms_val = nums[2]
            elif len(nums) == 2:
                icms_bc  = nums[0]
                icms_val = nums[1]
            elif nums:
                icms_bc = nums[0]
            break
    return icms_bc, icms_val


def extrair_pdf_cte(caminho, callback_progresso=None):
    with pdfplumber.open(caminho) as pdf:
        total_pags = len(pdf.pages)
        linhas = []
        for n, page in enumerate(pdf.pages):
            texto = page.extract_text()
            if texto:
                linhas.extend(texto.split("\n"))
            if callback_progresso:
                callback_progresso(int((n + 1) / total_pags * 50))

    registros = []
    i = 0
    while i < len(linhas):
        linha = linhas[i].strip()
        m = DOC_RE_CTE.match(linha)
        if m:
            fornec = None
            for j in range(i + 1, min(i + 6, len(linhas))):
                mf = FORN_RE.search(linhas[j])
                if mf:
                    fornec = mf
                    break
            cnpj_clean = limpar_cnpj(fornec.group(2)) if fornec else ""
            icms_bc, icms_val = parse_icms_pdf_cte(m.group(9) or "")
            registros.append({
                "data_entrada": m.group(1),
                "tipo":         m.group(2),
                "serie":        m.group(3),
                "numero":       m.group(4),
                "data_doc":     m.group(5),
                "uf":           m.group(7),
                "valor":        parse_valor_br(m.group(8)),
                "fornecedor":   fornec.group(1) if fornec else "",
                "cnpj":         cnpj_clean,
                "icms_bc":      icms_bc,
                "icms_val":     icms_val,
            })
        i += 1
    return registros


# ---------------------------------------------------------------------------
# Comparação — NF-e
# ---------------------------------------------------------------------------
def comparar_nfe(excel_rows, pdf_rows, callback_progresso=None):
    pdf_used = set()
    matched  = []
    so_excel = []
    total = len(excel_rows)
    for ei, e in enumerate(excel_rows):
        found = False
        for pi, p in enumerate(pdf_rows):
            if pi in pdf_used:
                continue
            cnpj_ok  = (e["cnpj"][:8] == p["cnpj"][:8]) if (e["cnpj"] and p["cnpj"]) else False
            num_ok   = (e["numero"] == p["numero"])
            valor_ok = abs(e["valor"] - p["valor"]) < 0.02
            if cnpj_ok and num_ok and valor_ok:
                matched.append({"excel": e, "pdf": p})
                pdf_used.add(pi)
                found = True
                break
        if not found:
            so_excel.append(e)
        if callback_progresso:
            callback_progresso(50 + int((ei + 1) / total * 50))
    so_pdf = [p for pi, p in enumerate(pdf_rows) if pi not in pdf_used]
    return matched, so_excel, so_pdf


# ---------------------------------------------------------------------------
# Comparação — CTe
# ---------------------------------------------------------------------------
def comparar_cte(excel_rows, pdf_rows, callback_progresso=None):
    pdf_used = set()
    matched  = []
    so_excel = []
    total = len(excel_rows)
    for ei, e in enumerate(excel_rows):
        found = False
        for pi, p in enumerate(pdf_rows):
            if pi in pdf_used:
                continue
            cnpj_e   = e["cnpj_emit"] or e["cnpj_tom"]
            cnpj_p   = p["cnpj"]
            cnpj_ok  = (cnpj_e[:8] == cnpj_p[:8]) if (cnpj_e and cnpj_p) else False
            num_ok   = (e["numero"] == p["numero"])
            valor_ok = abs(e["valor"] - p["valor"]) < 0.02
            if cnpj_ok and num_ok and valor_ok:
                matched.append({"excel": e, "pdf": p})
                pdf_used.add(pi)
                found = True
                break
        if not found:
            so_excel.append(e)
        if callback_progresso:
            callback_progresso(50 + int((ei + 1) / total * 50))
    so_pdf = [p for pi, p in enumerate(pdf_rows) if pi not in pdf_used]
    return matched, so_excel, so_pdf


# ---------------------------------------------------------------------------
# Comparação Teste — Teste.xls × RPT (xls/xlsx/csv)
# Teste.xls: colunas "Nr. documento" e "Valor contábil"
# RPT:       colunas SITUACAO (idx 5), NÚMERO_CTE (idx 24) e VALOR_TOTAL_PREST (idx 42)
# ---------------------------------------------------------------------------
def _ler_arquivo_generico(caminho):
    """Lê xls, xlsx ou csv e retorna um DataFrame, tentando vários métodos."""
    ext = os.path.splitext(caminho)[1].lower()

    if ext == ".csv":
        for enc in ("utf-8-sig", "latin1", "cp1252"):
            try:
                return pd.read_csv(caminho, encoding=enc, sep=None, engine="python")
            except Exception:
                continue
        raise ValueError(f"Não foi possível ler o arquivo CSV: {caminho}")

    if ext == ".xlsx" or ext == ".xlsm":
        return pd.read_excel(caminho, engine="openpyxl")

    if ext == ".xls":
        # 1) Tenta xlrd (XLS binário legítimo)
        if _TEM_XLRD:
            try:
                return pd.read_excel(caminho, engine="xlrd")
            except Exception:
                pass

        # 2) Tenta python-calamine (lê xls/xlsx/ods sem dependências extras)
        if _TEM_CALAMINE:
            try:
                return pd.read_excel(caminho, engine="calamine")
            except Exception:
                pass

        # 3) Fallback: HTML-XLS (gerado por sistemas como SAP/ERP/EFD)
        try:
            dfs = pd.read_html(caminho, decimal=",", thousands=".")
        except Exception:
            dfs = None
        if not dfs:
            for enc in ("latin1", "cp1252", "utf-8"):
                try:
                    dfs = pd.read_html(caminho, encoding=enc, decimal=",", thousands=".")
                    if dfs:
                        break
                except Exception:
                    continue
        if not dfs:
            raise ValueError(
                "Não foi possível ler o arquivo XLS.\n"
                "Instale 'xlrd' (pip install xlrd) para abrir arquivos .xls binários,\n"
                "ou verifique se o arquivo não está corrompido."
            )
        df = dfs[0]
        if df.iloc[0].apply(lambda x: isinstance(x, str)).all():
            df.columns = df.iloc[0]
            df = df.drop(0).reset_index(drop=True)
        return df

    # Qualquer outra extensão: tenta como Excel genérico
    return pd.read_excel(caminho, engine="openpyxl")


def extrair_teste(caminho):
    """Lê Teste.xls: colunas 'Nr. documento' e 'Valor contábil'."""
    import pandas as pd
    df = _ler_arquivo_generico(caminho)
    # Normaliza nomes de coluna
    df.columns = [str(c).strip() for c in df.columns]
    col_num = next((c for c in df.columns if "nr" in c.lower() or "número" in c.lower() or "numero" in c.lower() or "documento" in c.lower()), None)
    col_val = next((c for c in df.columns if "valor" in c.lower() or "contábil" in c.lower() or "contabil" in c.lower()), None)
    if col_num is None or col_val is None:
        raise ValueError(f"Colunas esperadas não encontradas no Teste.\nColunas encontradas: {list(df.columns)}")
    registros = []
    for _, row in df.iterrows():
        try:
            num = parse_num_documento(row[col_num])
            if not num:
                continue
            val = parse_valor_br(row[col_val])
            registros.append({"numero": num, "valor": val})
        except Exception:
            continue
    return registros


def extrair_rpt(caminho):
    """Lê RPT (xls/xlsx/csv): usa as colunas F (5), Y (24) e AQ (42) —
    SITUACAO, NÚMERO_CTE e VALOR_TOTAL_PREST."""
    df = _ler_arquivo_generico(caminho)
    df.columns = [str(c).strip() for c in df.columns]

    def _encontrar_col_num(cols):
        for c in cols:
            cu = c.upper().replace(" ", "_")
            if "NUMERO_CTE" in cu or "NÚMERO_CTE" in cu:
                return c
            if "N" in cu and "CTE" in cu:
                return c
        return None

    def _encontrar_col_val(cols):
        for c in cols:
            cu = c.upper().replace(" ", "_")
            if "VALOR_TOTAL_PREST" in cu or "VALOR_TOTAL" in cu:
                return c
        return None

    def _encontrar_col_situacao(cols):
        for c in cols:
            cu = c.upper().replace(" ", "_")
            if "SITUACAO" in cu or "SITUAÇÃO" in cu:
                return c
        return None

    # Se as colunas são numéricas (0,1,2...), a linha 0 tem os headers reais
    if df.columns.tolist() and all(isinstance(c, (int, float)) for c in df.columns[:3]):
        real_headers = [str(h).strip() for h in df.iloc[0].tolist()]
        df.columns = real_headers
        df = df.drop(0).reset_index(drop=True)

    col_num = None
    col_val = None
    col_sit = None

    # 1) Prioriza posição: F = índice 5 (SITUACAO), Y = índice 24 (NÚMERO_CTE), AQ = índice 42 (VALOR_TOTAL_PREST)
    IDX_SIT, IDX_NUM, IDX_VAL = 5, 24, 42
    if len(df.columns) > max(IDX_SIT, IDX_NUM, IDX_VAL):
        col_sit = df.columns[IDX_SIT]
        col_num = df.columns[IDX_NUM]
        col_val = df.columns[IDX_VAL]

    # 2) Fallback: busca por nome de coluna
    if col_num is None or col_val is None:
        col_num = _encontrar_col_num(df.columns)
        col_val = _encontrar_col_val(df.columns)
    if col_sit is None:
        col_sit = _encontrar_col_situacao(df.columns)

    if col_num is None or col_val is None:
        raise ValueError(f"Colunas esperadas não encontradas no RPT.\nColunas encontradas: {list(df.columns[:30])}")

    registros = []
    for _, row in df.iterrows():
        try:
            num = parse_num_documento(row[col_num])
            if not num:
                continue
            val = parse_valor_br(row[col_val])
            sit = str(row[col_sit]).strip() if (col_sit is not None and row[col_sit] is not None) else ""
            if sit.lower() in ("nan", "none"):
                sit = ""
            registros.append({"numero": num, "valor": val, "situacao": sit})
        except Exception:
            continue
    return registros


def comparar_teste_rpt(teste_rows, rpt_rows):
    """Compara por NÚMERO_CTE. Retorna (batidos, so_teste, so_rpt, divergentes)."""
    rpt_map = {}
    for r in rpt_rows:
        rpt_map.setdefault(r["numero"], []).append(r)

    batidos     = []  # numero, val_teste, val_rpt, situacao — valores iguais
    divergentes = []  # numero, val_teste, val_rpt, situacao — encontrados mas valores diferentes
    so_teste    = []  # presentes só no Teste

    teste_usados = set()
    for t in teste_rows:
        num = t["numero"]
        if num in rpt_map:
            rpt_regs = rpt_map[num]
            # Verifica se algum valor bate (tolerância de 0.02)
            matched_reg = None
            for rr in rpt_regs:
                if abs(t["valor"] - rr["valor"]) < 0.02:
                    matched_reg = rr
                    break
            if matched_reg is not None:
                batidos.append({
                    "numero": num, "val_teste": t["valor"], "val_rpt": matched_reg["valor"],
                    "situacao": matched_reg.get("situacao", ""),
                })
            else:
                divergentes.append({
                    "numero": num, "val_teste": t["valor"], "val_rpt": rpt_regs[0]["valor"],
                    "situacao": rpt_regs[0].get("situacao", ""),
                })
            teste_usados.add(num)
        else:
            so_teste.append({"numero": num, "valor": t["valor"], "situacao": ""})

    so_rpt = [
        {"numero": r["numero"], "valor": r["valor"], "situacao": r.get("situacao", "")}
        for r in rpt_rows if r["numero"] not in teste_usados
    ]
    return batidos, so_teste, so_rpt, divergentes


# ---------------------------------------------------------------------------
# Cores
# ---------------------------------------------------------------------------
VERDE    = "#2ecc71"
VERMELHO = "#e74c3c"
LARANJA  = "#e67e22"
AZUL     = "#2980b9"
AZUL_ESC = "#1a5276"
CINZA_BG = "#f4f6f9"
BRANCO   = "#ffffff"
TITULO_FG= "#2c3e50"
BTN_FG   = BRANCO


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Conferência Fiscal — SAT × Registro de Entradas (PDF)")
        self.geometry("1280x820")
        self.minsize(960, 640)
        self.configure(bg=CINZA_BG)
        self.resizable(True, True)

        self._modo = tk.StringVar(value="nfe")   # "nfe" ou "cte"
        self._excel_path = tk.StringVar()
        self._pdf_path   = tk.StringVar()
        self._status_msg = tk.StringVar(value="Selecione o modo, os arquivos e clique em Comparar.")
        self._resultado  = None

        # Comparação Teste
        self._teste_path = tk.StringVar()
        self._rpt_path   = tk.StringVar()
        self._resultado_teste = None

        self._build_ui()

    # ------------------------------------------------------------------
    def _build_ui(self):
        # ── Cabeçalho ──────────────────────────────────────────────────
        self._hdr = tk.Frame(self, bg=AZUL, pady=10)
        self._hdr.pack(fill="x")
        self._lbl_titulo = tk.Label(
            self._hdr,
            text="🔍  Conferência de Entradas Fiscais — Modelo 55 (NF-e)",
            font=("Helvetica", 15, "bold"), bg=AZUL, fg=BRANCO)
        self._lbl_titulo.pack()
        self._lbl_sub = tk.Label(
            self._hdr,
            text="SAT (Excel)  ×  Registro de Entradas (PDF)",
            font=("Helvetica", 10), bg=AZUL, fg="#d6eaf8")
        self._lbl_sub.pack()

        # ── Seletor de modo ────────────────────────────────────────────
        modo_frame = tk.Frame(self, bg=CINZA_BG, pady=6)
        modo_frame.pack(fill="x", padx=16)

        tk.Label(modo_frame, text="Tipo de documento:", bg=CINZA_BG,
                 font=("Helvetica", 10, "bold"), fg=TITULO_FG).pack(side="left")

        self._btn_nfe = tk.Button(
            modo_frame, text="NF-e (Modelo 55)",
            font=("Helvetica", 10, "bold"),
            bg=AZUL, fg=BRANCO, relief="flat", padx=16, pady=4,
            cursor="hand2", command=lambda: self._trocar_modo("nfe"))
        self._btn_nfe.pack(side="left", padx=(10, 4))

        self._btn_cte = tk.Button(
            modo_frame, text="CTe",
            font=("Helvetica", 10, "bold"),
            bg="#bdc3c7", fg="#555", relief="flat", padx=16, pady=4,
            cursor="hand2", command=lambda: self._trocar_modo("cte"))
        self._btn_cte.pack(side="left", padx=(0, 4))

        self._btn_teste = tk.Button(
            modo_frame, text="📋  Comparação Teste",
            font=("Helvetica", 10, "bold"),
            bg="#bdc3c7", fg="#555", relief="flat", padx=16, pady=4,
            cursor="hand2", command=lambda: self._trocar_modo("teste"))
        self._btn_teste.pack(side="left", padx=(0, 4))

        # ── Seleção de arquivos ────────────────────────────────────────
        self._sel = tk.LabelFrame(self, text="  Arquivos  ", bg=CINZA_BG,
                            font=("Helvetica", 10, "bold"), fg=TITULO_FG,
                            padx=10, pady=8)
        sel = self._sel
        sel.pack(fill="x", padx=16, pady=(4, 0))

        self._lbl_excel = tk.Label(sel, text="📊  Excel SAT (Modelo 55):",
                                   bg=CINZA_BG, width=26, anchor="w",
                                   font=("Helvetica", 10))
        self._lbl_excel.grid(row=0, column=0, sticky="w", pady=4)
        tk.Entry(sel, textvariable=self._excel_path, width=68,
                 font=("Helvetica", 9)).grid(row=0, column=1, sticky="ew", padx=(0, 8))
        tk.Button(sel, text="Procurar…",
                  command=lambda: self._browse(self._excel_path,
                                               [("Excel", ("*.xlsx", "*.xls"))]),
                  bg="#bdc3c7", relief="flat", padx=6,
                  cursor="hand2").grid(row=0, column=2)

        tk.Label(sel, text="📄  PDF (Registro Entradas):",
                 bg=CINZA_BG, width=26, anchor="w",
                 font=("Helvetica", 10)).grid(row=1, column=0, sticky="w", pady=4)
        tk.Entry(sel, textvariable=self._pdf_path, width=68,
                 font=("Helvetica", 9)).grid(row=1, column=1, sticky="ew", padx=(0, 8))
        tk.Button(sel, text="Procurar…",
                  command=lambda: self._browse(self._pdf_path,
                                               [("PDF", "*.pdf")]),
                  bg="#bdc3c7", relief="flat", padx=6,
                  cursor="hand2").grid(row=1, column=2)
        sel.columnconfigure(1, weight=1)

        # ── Painel exclusivo: Comparação Teste ─────────────────────────
        self._sel_teste = tk.LabelFrame(self, text="  Comparação Teste  ", bg=CINZA_BG,
                                        font=("Helvetica", 10, "bold"), fg=TITULO_FG,
                                        padx=10, pady=8)

        tk.Label(self._sel_teste, text="📊  Arquivo Teste (xls/xlsx/csv):",
                 bg=CINZA_BG, width=30, anchor="w",
                 font=("Helvetica", 10)).grid(row=0, column=0, sticky="w", pady=4)
        tk.Entry(self._sel_teste, textvariable=self._teste_path, width=64,
                 font=("Helvetica", 9)).grid(row=0, column=1, sticky="ew", padx=(0, 8))
        tk.Button(self._sel_teste, text="Procurar…",
                  command=lambda: self._browse(self._teste_path,
                                               [("Planilha", ("*.xls", "*.xlsx", "*.csv"))]),
                  bg="#bdc3c7", relief="flat", padx=6,
                  cursor="hand2").grid(row=0, column=2)

        tk.Label(self._sel_teste, text="📄  Arquivo RPT (xls/xlsx/csv):",
                 bg=CINZA_BG, width=30, anchor="w",
                 font=("Helvetica", 10)).grid(row=1, column=0, sticky="w", pady=4)
        tk.Entry(self._sel_teste, textvariable=self._rpt_path, width=64,
                 font=("Helvetica", 9)).grid(row=1, column=1, sticky="ew", padx=(0, 8))
        tk.Button(self._sel_teste, text="Procurar…",
                  command=lambda: self._browse(self._rpt_path,
                                               [("Planilha", ("*.xls", "*.xlsx", "*.csv"))]),
                  bg="#bdc3c7", relief="flat", padx=6,
                  cursor="hand2").grid(row=1, column=2)
        self._sel_teste.columnconfigure(1, weight=1)
        act = tk.Frame(self, bg=CINZA_BG)
        act.pack(fill="x", padx=16, pady=8)

        self._btn_comparar = tk.Button(
            act, text="▶  Comparar", font=("Helvetica", 11, "bold"),
            bg=AZUL, fg=BTN_FG, relief="flat", padx=18, pady=6,
            cursor="hand2", command=self._iniciar_comparacao)
        self._btn_comparar.pack(side="left")

        self._btn_exportar = tk.Button(
            act, text="💾  Exportar Excel", font=("Helvetica", 11, "bold"),
            bg="#27ae60", fg=BTN_FG, relief="flat", padx=18, pady=6,
            cursor="hand2", command=self._exportar_excel, state="disabled")
        self._btn_exportar.pack(side="left", padx=(10, 0))

        self._btn_exportar_teste = tk.Button(
            act, text="💾  Exportar Teste", font=("Helvetica", 11, "bold"),
            bg="#8e44ad", fg=BTN_FG, relief="flat", padx=18, pady=6,
            cursor="hand2", command=self._exportar_teste, state="disabled")
        self._btn_exportar_teste.pack(side="left", padx=(10, 0))

        self._progress = ttk.Progressbar(act, length=300, mode="determinate")
        self._progress.pack(side="left", padx=(20, 0))

        tk.Label(act, textvariable=self._status_msg, bg=CINZA_BG,
                 fg="#555", font=("Helvetica", 9)).pack(side="left", padx=12)

        # ── Notebook ───────────────────────────────────────────────────
        self._nb = ttk.Notebook(self)
        self._nb.pack(fill="both", expand=True, padx=16, pady=(0, 10))

        self._tab_match = self._criar_aba("✅  Encontrados nos dois", VERDE)
        self._tab_excel = self._criar_aba("⚠️  Só no Excel (SAT)",   VERMELHO)
        self._tab_pdf   = self._criar_aba("ℹ️  Só no PDF",            LARANJA)

        # Abas exclusivas do modo Comparação Teste (criadas uma vez, adicionadas/removidas dinamicamente)
        self._tab_batidos     = self._criar_aba("✅  Valores que batem",       VERDE)
        self._tab_divergentes = self._criar_aba("🔴  Valores divergentes",     VERMELHO)
        self._tab_so_teste    = self._criar_aba("⚠️  Só no Teste",             LARANJA)
        self._tab_so_rpt      = self._criar_aba("ℹ️  Só no RPT",              AZUL)

        # Inicialmente esconde as abas do modo Teste
        for tab in (self._tab_batidos, self._tab_divergentes,
                    self._tab_so_teste, self._tab_so_rpt):
            self._nb.hide(tab)

        # ── Rodapé ─────────────────────────────────────────────────────
        rod = tk.Frame(self, bg="#dfe6e9", pady=4)
        rod.pack(fill="x", side="bottom")
        self._lbl_totais = tk.Label(rod, text="", bg="#dfe6e9",
                                    fg=TITULO_FG, font=("Helvetica", 9, "bold"))
        self._lbl_totais.pack()

    # ------------------------------------------------------------------
    def _trocar_modo(self, modo):
        if self._modo.get() == modo:
            return
        self._modo.set(modo)
        self._resultado = None
        self._resultado_teste = None
        self._btn_exportar.config(state="disabled")
        self._btn_exportar_teste.config(state="disabled")
        self._lbl_totais.config(text="")

        # Limpa abas padrão
        for tab in (self._tab_match, self._tab_excel, self._tab_pdf):
            tab._tv.delete(*tab._tv.get_children())
            tab._lbl_count.config(text="0 registros")
        # Limpa abas Teste
        for tab in (self._tab_batidos, self._tab_divergentes,
                    self._tab_so_teste, self._tab_so_rpt):
            tab._tv.delete(*tab._tv.get_children())
            tab._lbl_count.config(text="0 registros")

        if modo == "teste":
            # Mostra painel Teste, esconde painel normal
            self._sel.pack_forget()
            self._sel_teste.pack(fill="x", padx=16, pady=(4, 0))
            # Esconde abas normais, mostra abas Teste
            for tab in (self._tab_match, self._tab_excel, self._tab_pdf):
                self._nb.hide(tab)
            for tab in (self._tab_batidos, self._tab_divergentes,
                        self._tab_so_teste, self._tab_so_rpt):
                self._nb.add(tab)
            self._btn_teste.config(bg=AZUL, fg=BRANCO)
            self._btn_nfe.config(bg="#bdc3c7", fg="#555")
            self._btn_cte.config(bg="#bdc3c7", fg="#555")
            self._lbl_titulo.config(text="📋  Comparação Teste × RPT")
            self._btn_comparar.config(
                text="▶  Comparar Teste",
                command=self._iniciar_comparacao_teste)
        else:
            # Esconde painel Teste, mostra painel normal
            self._sel_teste.pack_forget()
            self._sel.pack(fill="x", padx=16, pady=(4, 0))
            # Mostra abas normais, esconde abas Teste
            for tab in (self._tab_match, self._tab_excel, self._tab_pdf):
                self._nb.add(tab)
            for tab in (self._tab_batidos, self._tab_divergentes,
                        self._tab_so_teste, self._tab_so_rpt):
                self._nb.hide(tab)
            self._btn_comparar.config(
                text="▶  Comparar",
                command=self._iniciar_comparacao)
            if modo == "nfe":
                self._btn_nfe.config(bg=AZUL, fg=BRANCO)
                self._btn_cte.config(bg="#bdc3c7", fg="#555")
                self._btn_teste.config(bg="#bdc3c7", fg="#555")
                self._lbl_titulo.config(
                    text="🔍  Conferência de Entradas Fiscais — Modelo 55 (NF-e)")
                self._lbl_excel.config(text="📊  Excel SAT (Modelo 55):")
            else:
                self._btn_cte.config(bg=AZUL, fg=BRANCO)
                self._btn_nfe.config(bg="#bdc3c7", fg="#555")
                self._btn_teste.config(bg="#bdc3c7", fg="#555")
                self._lbl_titulo.config(
                    text="🔍  Conferência de CTe's Fiscais")
                self._lbl_excel.config(text="📊  Excel SAT (CTe):")

        self._status_msg.set("Selecione os arquivos e clique em Comparar.")

    # ------------------------------------------------------------------
    def _browse(self, var, ftypes):
        path = filedialog.askopenfilename(filetypes=ftypes + [("Todos", "*.*")])
        if path:
            var.set(path)

    # ------------------------------------------------------------------
    def _criar_aba(self, titulo, cor):
        frame = tk.Frame(self._nb, bg=BRANCO)
        self._nb.add(frame, text=titulo)

        hdr = tk.Frame(frame, bg=cor, pady=4)
        hdr.pack(fill="x")
        tk.Label(hdr, text=titulo, bg=cor, fg=BRANCO,
                 font=("Helvetica", 10, "bold")).pack(side="left", padx=10)
        lbl_count = tk.Label(hdr, text="0 registros", bg=cor, fg=BRANCO,
                             font=("Helvetica", 9))
        lbl_count.pack(side="right", padx=10)
        frame._lbl_count = lbl_count

        wrap = tk.Frame(frame, bg=BRANCO)
        wrap.pack(fill="both", expand=True)
        vsb = ttk.Scrollbar(wrap, orient="vertical")
        hsb = ttk.Scrollbar(wrap, orient="horizontal")
        tv  = ttk.Treeview(wrap, yscrollcommand=vsb.set,
                           xscrollcommand=hsb.set, selectmode="browse")
        vsb.config(command=tv.yview)
        hsb.config(command=tv.xview)
        vsb.pack(side="right",  fill="y")
        hsb.pack(side="bottom", fill="x")
        tv.pack(fill="both", expand=True)
        frame._tv = tv
        return frame

    # ------------------------------------------------------------------
    def _iniciar_comparacao(self):
        excel = self._excel_path.get().strip()
        pdf   = self._pdf_path.get().strip()
        if not excel or not os.path.isfile(excel):
            messagebox.showwarning("Arquivo faltando", "Selecione o arquivo Excel SAT.")
            return
        if not pdf or not os.path.isfile(pdf):
            messagebox.showwarning("Arquivo faltando", "Selecione o arquivo PDF.")
            return
        self._btn_comparar.config(state="disabled")
        self._btn_exportar.config(state="disabled")
        self._progress["value"] = 0
        self._status_msg.set("Processando…")
        threading.Thread(target=self._executar_comparacao,
                         args=(excel, pdf), daemon=True).start()

    def _executar_comparacao(self, excel, pdf):
        try:
            modo = self._modo.get()
            self._set_status("Lendo Excel…", 5)

            if modo == "nfe":
                excel_rows = extrair_excel_nfe(excel)
                self._set_status("Lendo PDF…", 10)
                pdf_rows = extrair_pdf_nfe(pdf,
                    callback_progresso=lambda v: self._set_progress(10 + v // 2))
                self._set_status("Comparando…", 60)
                matched, so_excel, so_pdf = comparar_nfe(excel_rows, pdf_rows,
                    callback_progresso=lambda v: self._set_progress(60 + v // 2))
            else:
                excel_rows = extrair_excel_cte(excel)
                self._set_status("Lendo PDF…", 10)
                pdf_rows = extrair_pdf_cte(pdf,
                    callback_progresso=lambda v: self._set_progress(10 + v // 2))
                self._set_status("Comparando…", 60)
                matched, so_excel, so_pdf = comparar_cte(excel_rows, pdf_rows,
                    callback_progresso=lambda v: self._set_progress(60 + v // 2))

            self._resultado = (matched, so_excel, so_pdf, modo)
            self.after(0, lambda: self._mostrar_resultado(matched, so_excel, so_pdf, modo))
        except Exception as ex:
            self.after(0, lambda: messagebox.showerror("Erro", str(ex)))
            self.after(0, lambda: self._btn_comparar.config(state="normal"))
            self.after(0, lambda: self._status_msg.set("Erro durante o processamento."))

    def _set_status(self, msg, prog):
        self.after(0, lambda: self._status_msg.set(msg))
        self.after(0, lambda: self._set_progress(prog))

    def _set_progress(self, val):
        self._progress["value"] = min(val, 100)

    # ------------------------------------------------------------------
    def _mostrar_resultado(self, matched, so_excel, so_pdf, modo):
        if modo == "nfe":
            self._preencher_match_nfe(self._tab_match, matched)
            self._preencher_excel_nfe(self._tab_excel, so_excel)
            self._preencher_pdf_nfe(self._tab_pdf, so_pdf)
        else:
            self._preencher_match_cte(self._tab_match, matched)
            self._preencher_excel_cte(self._tab_excel, so_excel)
            self._preencher_pdf_cte(self._tab_pdf, so_pdf)

        self._tab_match._lbl_count.config(text=f"{len(matched)} registros")
        self._tab_excel._lbl_count.config(text=f"{len(so_excel)} registros")
        self._tab_pdf._lbl_count.config(text=f"{len(so_pdf)} registros")

        val_match = sum(r["excel"]["valor"] for r in matched)
        val_excel = sum(r["valor"] for r in so_excel)
        self._lbl_totais.config(text=(
            f"✅ {len(matched)} encontrados  |  "
            f"⚠️ {len(so_excel)} só no Excel  |  "
            f"ℹ️ {len(so_pdf)} só no PDF  ||  "
            f"Valor conferido: R$ {val_match:,.2f}  |  "
            f"Valor sem par no PDF: R$ {val_excel:,.2f}"
        ))
        self._progress["value"] = 100
        self._status_msg.set("Concluído!")
        self._btn_comparar.config(state="normal")
        self._btn_exportar.config(state="normal")
        self._nb.select(0)

    # ------------------------------------------------------------------
    # Preenchimento NF-e
    # ------------------------------------------------------------------
    def _preencher_match_nfe(self, frame, dados):
        tv = frame._tv
        tv.delete(*tv.get_children())
        colunas = (
            "numero", "serie", "nome", "cnpj", "uf",
            "valor_excel", "valor_pdf", "dif_valor",
            "icms_bc_excel", "icms_bc_pdf",
            "icms_val_excel", "icms_val_pdf",
            "ipi_excel", "ipi_pdf",
            "tipo_op", "situacao", "data_emissao", "data_entrada_pdf",
        )
        tv["columns"] = colunas
        tv["show"]    = "headings"
        cab = {
            "numero":           ("Número NF",       90),
            "serie":            ("Série",            50),
            "nome":             ("Emitente",        200),
            "cnpj":             ("CNPJ",            130),
            "uf":               ("UF",               42),
            "valor_excel":      ("Valor Excel",     105),
            "valor_pdf":        ("Valor PDF",       105),
            "dif_valor":        ("Dif. Valor",       95),
            "icms_bc_excel":    ("BC ICMS Excel",  115),
            "icms_bc_pdf":      ("BC ICMS PDF",    115),
            "icms_val_excel":   ("ICMS Excel",      100),
            "icms_val_pdf":     ("ICMS PDF",        100),
            "ipi_excel":        ("IPI Excel",        90),
            "ipi_pdf":          ("IPI PDF",          90),
            "tipo_op":          ("E/S",              42),
            "situacao":         ("Situação",          90),
            "data_emissao":     ("Dt Emissão",        90),
            "data_entrada_pdf": ("Dt Entrada PDF",  110),
        }
        for col, (txt, w) in cab.items():
            tv.heading(col, text=txt, command=lambda c=col: self._ordenar(tv, c, False))
            tv.column(col, width=w, anchor="center")
        tv.column("nome", anchor="w")
        tv.tag_configure("dif_pos", background="#FFF3CD")
        tv.tag_configure("dif_neg", background="#FADBD8")
        for r in dados:
            e, p = r["excel"], r["pdf"]
            dif = p["valor"] - e["valor"]
            if abs(dif) < 0.02:
                dif_str, tag = "—", ()
            elif dif > 0:
                dif_str, tag = f"+R$ {dif:,.2f}", ("dif_pos",)
            else:
                dif_str, tag = f"-R$ {abs(dif):,.2f}", ("dif_neg",)
            tv.insert("", "end", tags=tag, values=(
                e["numero"], e["serie"], e["nome"], e["cnpj"], e["uf"] or p["uf"],
                f'R$ {e["valor"]:,.2f}', f'R$ {p["valor"]:,.2f}', dif_str,
                f'R$ {e["icms_bc"]:,.2f}', f'R$ {p["icms_bc"]:,.2f}',
                f'R$ {e["icms_val"]:,.2f}', f'R$ {p["icms_val"]:,.2f}',
                f'R$ {e["ipi_val"]:,.2f}', f'R$ {p["ipi_val"]:,.2f}',
                e["tipo_op"], e["situacao"], e["data_emissao"], p["data_entrada"],
            ))

    def _preencher_excel_nfe(self, frame, dados):
        tv = frame._tv
        tv.delete(*tv.get_children())
        colunas = ("numero", "serie", "nome", "cnpj", "uf",
                   "valor", "icms_bc", "icms_val", "ipi_val",
                   "tipo_op", "situacao", "data_emissao", "modelo")
        tv["columns"] = colunas
        tv["show"]    = "headings"
        cab = {
            "numero":       ("Número NF",    90),
            "serie":        ("Série",         50),
            "nome":         ("Emitente",     200),
            "cnpj":         ("CNPJ",         130),
            "uf":           ("UF",            42),
            "valor":        ("Valor",        105),
            "icms_bc":      ("BC ICMS",      110),
            "icms_val":     ("ICMS",         100),
            "ipi_val":      ("IPI",           90),
            "tipo_op":      ("E/S",           42),
            "situacao":     ("Situação",       90),
            "data_emissao": ("Dt Emissão",    90),
            "modelo":       ("Modelo",         60),
        }
        for col, (txt, w) in cab.items():
            tv.heading(col, text=txt, command=lambda c=col: self._ordenar(tv, c, False))
            tv.column(col, width=w, anchor="center")
        tv.column("nome", anchor="w")
        tv.tag_configure("cancelado", foreground="#aaa")
        for e in dados:
            tag = ("cancelado",) if e["situacao"] == "Cancelado" else ()
            tv.insert("", "end", tags=tag, values=(
                e["numero"], e["serie"], e["nome"], e["cnpj"], e["uf"],
                f'R$ {e["valor"]:,.2f}',
                f'R$ {e["icms_bc"]:,.2f}',
                f'R$ {e["icms_val"]:,.2f}',
                f'R$ {e["ipi_val"]:,.2f}',
                e["tipo_op"], e["situacao"], e["data_emissao"], e["modelo"],
            ))

    def _preencher_pdf_nfe(self, frame, dados):
        tv = frame._tv
        tv.delete(*tv.get_children())
        colunas = ("numero", "serie", "tipo", "fornecedor", "cnpj",
                   "uf", "valor", "icms_bc", "icms_val", "ipi_val",
                   "data_entrada", "data_doc")
        tv["columns"] = colunas
        tv["show"]    = "headings"
        cab = {
            "numero":       ("Número NF",    90),
            "serie":        ("Série",         50),
            "tipo":         ("Tipo",          60),
            "fornecedor":   ("Fornecedor",   200),
            "cnpj":         ("CNPJ",         130),
            "uf":           ("UF",            42),
            "valor":        ("Valor",        105),
            "icms_bc":      ("BC ICMS",      110),
            "icms_val":     ("ICMS",         100),
            "ipi_val":      ("IPI",           90),
            "data_entrada": ("Dt Entrada",    90),
            "data_doc":     ("Dt Documento",  90),
        }
        for col, (txt, w) in cab.items():
            tv.heading(col, text=txt, command=lambda c=col: self._ordenar(tv, c, False))
            tv.column(col, width=w, anchor="center")
        tv.column("fornecedor", anchor="w")
        for p in dados:
            tv.insert("", "end", values=(
                p["numero"], p["serie"], p["tipo"],
                p["fornecedor"], p["cnpj"], p["uf"],
                f'R$ {p["valor"]:,.2f}',
                f'R$ {p["icms_bc"]:,.2f}',
                f'R$ {p["icms_val"]:,.2f}',
                f'R$ {p["ipi_val"]:,.2f}',
                p["data_entrada"], p["data_doc"],
            ))

    # ------------------------------------------------------------------
    # Preenchimento CTe
    # ------------------------------------------------------------------
    def _preencher_match_cte(self, frame, dados):
        tv = frame._tv
        tv.delete(*tv.get_children())
        colunas = (
            "numero", "serie", "nome_emit", "cnpj_emit", "papel",
            "valor_excel", "valor_pdf", "dif_valor",
            "icms_bc_excel", "icms_bc_pdf",
            "icms_val_excel", "icms_val_pdf",
            "situacao", "data_emissao", "data_entrada_pdf",
        )
        tv["columns"] = colunas
        tv["show"]    = "headings"
        cab = {
            "numero":           ("Número CTe",     100),
            "serie":            ("Série",           50),
            "nome_emit":        ("Emitente",       200),
            "cnpj_emit":        ("CNPJ Emitente",  130),
            "papel":            ("Papel Tomador",  110),
            "valor_excel":      ("Valor Excel",    105),
            "valor_pdf":        ("Valor PDF",      105),
            "dif_valor":        ("Dif. Valor",      95),
            "icms_bc_excel":    ("BC ICMS Excel",  115),
            "icms_bc_pdf":      ("BC ICMS PDF",    115),
            "icms_val_excel":   ("ICMS Excel",     100),
            "icms_val_pdf":     ("ICMS PDF",       100),
            "situacao":         ("Situação",         90),
            "data_emissao":     ("Dt Emissão",       90),
            "data_entrada_pdf": ("Dt Entrada PDF",  110),
        }
        for col, (txt, w) in cab.items():
            tv.heading(col, text=txt, command=lambda c=col: self._ordenar(tv, c, False))
            tv.column(col, width=w, anchor="center")
        tv.column("nome_emit", anchor="w")
        tv.tag_configure("dif_pos", background="#FFF3CD")
        tv.tag_configure("dif_neg", background="#FADBD8")
        for r in dados:
            e, p = r["excel"], r["pdf"]
            dif = p["valor"] - e["valor"]
            if abs(dif) < 0.02:
                dif_str, tag = "—", ()
            elif dif > 0:
                dif_str, tag = f"+R$ {dif:,.2f}", ("dif_pos",)
            else:
                dif_str, tag = f"-R$ {abs(dif):,.2f}", ("dif_neg",)
            tv.insert("", "end", tags=tag, values=(
                e["numero"], e["serie"], e["nome_emit"], e["cnpj_emit"], e["papel"],
                f'R$ {e["valor"]:,.2f}', f'R$ {p["valor"]:,.2f}', dif_str,
                f'R$ {e["icms_bc"]:,.2f}', f'R$ {p["icms_bc"]:,.2f}',
                f'R$ {e["icms_val"]:,.2f}', f'R$ {p["icms_val"]:,.2f}',
                e["situacao"], e["data_emissao"], p["data_entrada"],
            ))

    def _preencher_excel_cte(self, frame, dados):
        tv = frame._tv
        tv.delete(*tv.get_children())
        colunas = ("numero", "serie", "nome_emit", "cnpj_emit", "papel",
                   "valor", "icms_bc", "icms_val",
                   "situacao", "data_emissao")
        tv["columns"] = colunas
        tv["show"]    = "headings"
        cab = {
            "numero":       ("Número CTe",    100),
            "serie":        ("Série",          50),
            "nome_emit":    ("Emitente",      200),
            "cnpj_emit":    ("CNPJ Emitente", 130),
            "papel":        ("Papel Tomador", 110),
            "valor":        ("Valor",         105),
            "icms_bc":      ("BC ICMS",       110),
            "icms_val":     ("ICMS",          100),
            "situacao":     ("Situação",        90),
            "data_emissao": ("Dt Emissão",      90),
        }
        for col, (txt, w) in cab.items():
            tv.heading(col, text=txt, command=lambda c=col: self._ordenar(tv, c, False))
            tv.column(col, width=w, anchor="center")
        tv.column("nome_emit", anchor="w")
        tv.tag_configure("cancelado", foreground="#aaa")
        for e in dados:
            tag = ("cancelado",) if e["situacao"].upper() == "CANCELADO" else ()
            tv.insert("", "end", tags=tag, values=(
                e["numero"], e["serie"], e["nome_emit"], e["cnpj_emit"], e["papel"],
                f'R$ {e["valor"]:,.2f}',
                f'R$ {e["icms_bc"]:,.2f}',
                f'R$ {e["icms_val"]:,.2f}',
                e["situacao"], e["data_emissao"],
            ))

    def _preencher_pdf_cte(self, frame, dados):
        tv = frame._tv
        tv.delete(*tv.get_children())
        colunas = ("numero", "serie", "tipo", "fornecedor", "cnpj",
                   "uf", "valor", "icms_bc", "icms_val",
                   "data_entrada", "data_doc")
        tv["columns"] = colunas
        tv["show"]    = "headings"
        cab = {
            "numero":       ("Número CTe",   100),
            "serie":        ("Série",         50),
            "tipo":         ("Tipo",          55),
            "fornecedor":   ("Fornecedor",   200),
            "cnpj":         ("CNPJ",         130),
            "uf":           ("UF",            42),
            "valor":        ("Valor",        105),
            "icms_bc":      ("BC ICMS",      110),
            "icms_val":     ("ICMS",         100),
            "data_entrada": ("Dt Entrada",    90),
            "data_doc":     ("Dt Documento",  90),
        }
        for col, (txt, w) in cab.items():
            tv.heading(col, text=txt, command=lambda c=col: self._ordenar(tv, c, False))
            tv.column(col, width=w, anchor="center")
        tv.column("fornecedor", anchor="w")
        for p in dados:
            tv.insert("", "end", values=(
                p["numero"], p["serie"], p["tipo"],
                p["fornecedor"], p["cnpj"], p["uf"],
                f'R$ {p["valor"]:,.2f}',
                f'R$ {p["icms_bc"]:,.2f}',
                f'R$ {p["icms_val"]:,.2f}',
                p["data_entrada"], p["data_doc"],
            ))

    # ------------------------------------------------------------------
    def _ordenar(self, tv, col, reverse):
        items = [(tv.set(k, col), k) for k in tv.get_children("")]
        try:
            items.sort(
                key=lambda t: float(
                    t[0].replace("R$", "").replace(".", "")
                        .replace(",", ".").replace("+", "")
                        .replace("-", "").replace("—", "0").strip()),
                reverse=reverse)
        except ValueError:
            items.sort(reverse=reverse)
        for idx, (_, k) in enumerate(items):
            tv.move(k, "", idx)
        tv.heading(col, command=lambda: self._ordenar(tv, col, not reverse))

    # ------------------------------------------------------------------
    def _exportar_excel(self):
        if not self._resultado:
            return
        matched, so_excel, so_pdf, modo = self._resultado
        nome_default = (
            "conferencia_nfe_resultado.xlsx"
            if modo == "nfe" else
            "conferencia_cte_resultado.xlsx"
        )
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", ("*.xlsx", "*.xls")), ("Todos os arquivos", "*.*")],
            initialfile=nome_default)
        if not path:
            return
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            wb = openpyxl.Workbook()

            if modo == "nfe":
                ws1 = wb.active
                ws1.title = "Conferidos"
                ws1.append([
                    "Número NF", "Série", "Emitente", "CNPJ", "UF",
                    "Valor Excel", "Valor PDF", "Dif. Valor (PDF-Excel)",
                    "BC ICMS Excel", "BC ICMS PDF",
                    "ICMS Excel", "ICMS PDF",
                    "IPI Excel", "IPI PDF",
                    "E/S", "Situação", "Dt Emissão", "Dt Entrada PDF",
                ])
                for r in matched:
                    e, p = r["excel"], r["pdf"]
                    ws1.append([
                        e["numero"], e["serie"], e["nome"], e["cnpj"], e["uf"] or p["uf"],
                        e["valor"], p["valor"], round(p["valor"] - e["valor"], 2),
                        e["icms_bc"], p["icms_bc"],
                        e["icms_val"], p["icms_val"],
                        e["ipi_val"], p["ipi_val"],
                        e["tipo_op"], e["situacao"], e["data_emissao"], p["data_entrada"],
                    ])
                ws2 = wb.create_sheet("So no Excel")
                ws2.append([
                    "Número NF", "Série", "Emitente", "CNPJ", "UF",
                    "Valor", "BC ICMS", "ICMS", "IPI",
                    "E/S", "Situação", "Dt Emissão", "Modelo",
                ])
                for e in so_excel:
                    ws2.append([
                        e["numero"], e["serie"], e["nome"], e["cnpj"], e["uf"],
                        e["valor"], e["icms_bc"], e["icms_val"], e["ipi_val"],
                        e["tipo_op"], e["situacao"], e["data_emissao"], e["modelo"],
                    ])
                ws3 = wb.create_sheet("So no PDF")
                ws3.append([
                    "Número NF", "Série", "Tipo", "Fornecedor", "CNPJ", "UF",
                    "Valor", "BC ICMS", "ICMS", "IPI",
                    "Dt Entrada", "Dt Documento",
                ])
                for p in so_pdf:
                    ws3.append([
                        p["numero"], p["serie"], p["tipo"],
                        p["fornecedor"], p["cnpj"], p["uf"],
                        p["valor"], p["icms_bc"], p["icms_val"], p["ipi_val"],
                        p["data_entrada"], p["data_doc"],
                    ])
            else:
                ws1 = wb.active
                ws1.title = "Conferidos"
                ws1.append([
                    "Número CTe", "Série", "Emitente", "CNPJ Emitente", "Papel Tomador",
                    "Valor Excel", "Valor PDF", "Dif. Valor (PDF-Excel)",
                    "BC ICMS Excel", "BC ICMS PDF",
                    "ICMS Excel", "ICMS PDF",
                    "Situação", "Dt Emissão", "Dt Entrada PDF",
                ])
                for r in matched:
                    e, p = r["excel"], r["pdf"]
                    ws1.append([
                        e["numero"], e["serie"], e["nome_emit"], e["cnpj_emit"], e["papel"],
                        e["valor"], p["valor"], round(p["valor"] - e["valor"], 2),
                        e["icms_bc"], p["icms_bc"],
                        e["icms_val"], p["icms_val"],
                        e["situacao"], e["data_emissao"], p["data_entrada"],
                    ])
                ws2 = wb.create_sheet("So no Excel")
                ws2.append([
                    "Número CTe", "Série", "Emitente", "CNPJ Emitente", "Papel Tomador",
                    "Valor", "BC ICMS", "ICMS",
                    "Situação", "Dt Emissão",
                ])
                for e in so_excel:
                    ws2.append([
                        e["numero"], e["serie"], e["nome_emit"], e["cnpj_emit"], e["papel"],
                        e["valor"], e["icms_bc"], e["icms_val"],
                        e["situacao"], e["data_emissao"],
                    ])
                ws3 = wb.create_sheet("So no PDF")
                ws3.append([
                    "Número CTe", "Série", "Tipo", "Fornecedor", "CNPJ", "UF",
                    "Valor", "BC ICMS", "ICMS",
                    "Dt Entrada", "Dt Documento",
                ])
                for p in so_pdf:
                    ws3.append([
                        p["numero"], p["serie"], p["tipo"],
                        p["fornecedor"], p["cnpj"], p["uf"],
                        p["valor"], p["icms_bc"], p["icms_val"],
                        p["data_entrada"], p["data_doc"],
                    ])

            hfont  = Font(bold=True, color="FFFFFF")
            center = Alignment(horizontal="center")
            fills  = {ws1: "1B5E20", ws2: "B71C1C", ws3: "E65100"}
            for ws, cor in fills.items():
                fill = PatternFill("solid", fgColor=cor)
                for cell in ws[1]:
                    cell.fill = fill
                    cell.font = hfont
                    cell.alignment = center
                for col in ws.columns:
                    ml = max((len(str(c.value or "")) for c in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 52)

            wb.save(path)
            messagebox.showinfo("Exportado!", f"Resultado salvo em:\n{path}")
        except Exception as ex:
            messagebox.showerror("Erro ao exportar", str(ex))


    # ------------------------------------------------------------------
    # Comparação Teste × RPT
    # ------------------------------------------------------------------
    def _iniciar_comparacao_teste(self):
        teste = self._teste_path.get().strip()
        rpt   = self._rpt_path.get().strip()
        if not teste or not os.path.isfile(teste):
            messagebox.showwarning("Arquivo faltando", "Selecione o arquivo Teste.")
            return
        if not rpt or not os.path.isfile(rpt):
            messagebox.showwarning("Arquivo faltando", "Selecione o arquivo RPT.")
            return
        self._btn_comparar.config(state="disabled")
        self._btn_exportar_teste.config(state="disabled")
        self._progress["value"] = 0
        self._status_msg.set("Processando comparação Teste × RPT…")
        threading.Thread(target=self._executar_comparacao_teste,
                         args=(teste, rpt), daemon=True).start()

    def _executar_comparacao_teste(self, teste_path, rpt_path):
        try:
            self._set_status("Lendo arquivo Teste…", 10)
            teste_rows = extrair_teste(teste_path)
            self._set_status("Lendo arquivo RPT…", 40)
            rpt_rows = extrair_rpt(rpt_path)
            self._set_status("Comparando…", 70)
            batidos, so_teste, so_rpt, divergentes = comparar_teste_rpt(teste_rows, rpt_rows)
            self._resultado_teste = (batidos, so_teste, so_rpt, divergentes)
            self.after(0, lambda: self._mostrar_resultado_teste(batidos, so_teste, so_rpt, divergentes))
        except Exception as ex:
            self.after(0, lambda: messagebox.showerror("Erro", str(ex)))
            self.after(0, lambda: self._btn_comparar.config(state="normal"))
            self.after(0, lambda: self._status_msg.set("Erro durante o processamento."))

    def _mostrar_resultado_teste(self, batidos, so_teste, so_rpt, divergentes):
        self._preencher_tab_teste(self._tab_batidos, batidos,
                                  ("numero", "val_teste", "val_rpt", "situacao"),
                                  ("Número CTe", "Valor Teste", "Valor RPT", "Situação"))
        self._preencher_tab_teste(self._tab_divergentes, divergentes,
                                  ("numero", "val_teste", "val_rpt", "diferenca", "situacao"),
                                  ("Número CTe", "Valor Teste", "Valor RPT", "Diferença", "Situação"))
        self._preencher_tab_simples(self._tab_so_teste, so_teste,
                                    ("numero", "valor", "situacao"),
                                    ("Número CTe", "Valor Teste", "Situação"))
        self._preencher_tab_simples(self._tab_so_rpt, so_rpt,
                                    ("numero", "valor", "situacao"),
                                    ("Número CTe", "Valor RPT", "Situação"))

        self._tab_batidos._lbl_count.config(text=f"{len(batidos)} registros")
        self._tab_divergentes._lbl_count.config(text=f"{len(divergentes)} registros")
        self._tab_so_teste._lbl_count.config(text=f"{len(so_teste)} registros")
        self._tab_so_rpt._lbl_count.config(text=f"{len(so_rpt)} registros")

        self._lbl_totais.config(
            text=f"Batem: {len(batidos)}  |  Divergentes: {len(divergentes)}  |  "
                 f"Só no Teste: {len(so_teste)}  |  Só no RPT: {len(so_rpt)}")
        self._btn_exportar_teste.config(state="normal")
        self._btn_comparar.config(state="normal")
        self._set_progress(100)
        self._status_msg.set("Comparação Teste × RPT concluída.")

    def _preencher_tab_teste(self, frame, dados, cols, headers):
        tv = frame._tv
        tv.delete(*tv.get_children())
        tv["columns"] = cols
        tv["show"]    = "headings"
        widths = {"numero": 130, "val_teste": 130, "val_rpt": 130,
                  "diferenca": 130, "situacao": 130}
        for col, hdr in zip(cols, headers):
            tv.heading(col, text=hdr, command=lambda c=col: self._ordenar(tv, c, False))
            tv.column(col, width=widths.get(col, 120), anchor="center")
        for d in dados:
            if "diferenca" in cols:
                dif = round(d["val_teste"] - d["val_rpt"], 2)
                tv.insert("", "end", values=(
                    d["numero"],
                    f'R$ {d["val_teste"]:,.2f}',
                    f'R$ {d["val_rpt"]:,.2f}',
                    f'R$ {dif:,.2f}',
                    d.get("situacao", ""),
                ))
            else:
                tv.insert("", "end", values=(
                    d["numero"],
                    f'R$ {d["val_teste"]:,.2f}',
                    f'R$ {d["val_rpt"]:,.2f}',
                    d.get("situacao", ""),
                ))

    def _preencher_tab_simples(self, frame, dados, cols, headers):
        tv = frame._tv
        tv.delete(*tv.get_children())
        tv["columns"] = cols
        tv["show"]    = "headings"
        for col, hdr in zip(cols, headers):
            tv.heading(col, text=hdr, command=lambda c=col: self._ordenar(tv, c, False))
            tv.column(col, width=150, anchor="center")
        for d in dados:
            tv.insert("", "end", values=(
                d["numero"],
                f'R$ {d["valor"]:,.2f}',
                d.get("situacao", ""),
            ))

    def _exportar_teste(self):
        if not self._resultado_teste:
            return
        batidos, so_teste, so_rpt, divergentes = self._resultado_teste
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", ("*.xlsx", "*.xls")), ("Todos os arquivos", "*.*")],
            initialfile="comparacao_teste_resultado.xlsx")
        if not path:
            return
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            wb = openpyxl.Workbook()

            def _escrever_aba(ws, linhas, cabecalho, cor):
                ws.append(cabecalho)
                for row in linhas:
                    ws.append(row)
                fill  = PatternFill("solid", fgColor=cor)
                hfont = Font(bold=True, color="FFFFFF")
                center = Alignment(horizontal="center")
                for cell in ws[1]:
                    cell.fill  = fill
                    cell.font  = hfont
                    cell.alignment = center
                for col in ws.columns:
                    ml = max((len(str(c.value or "")) for c in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 40)

            ws1 = wb.active
            ws1.title = "Valores que batem"
            _escrever_aba(ws1,
                [[d["numero"], d["val_teste"], d["val_rpt"], d.get("situacao", "")] for d in batidos],
                ["Número CTe", "Valor Teste", "Valor RPT", "Situação"],
                "1B5E20")

            ws2 = wb.create_sheet("Valores divergentes")
            _escrever_aba(ws2,
                [[d["numero"], d["val_teste"], d["val_rpt"],
                  round(d["val_teste"] - d["val_rpt"], 2), d.get("situacao", "")] for d in divergentes],
                ["Número CTe", "Valor Teste", "Valor RPT", "Diferença", "Situação"],
                "B71C1C")

            ws3 = wb.create_sheet("Só no Teste")
            _escrever_aba(ws3,
                [[d["numero"], d["valor"], d.get("situacao", "")] for d in so_teste],
                ["Número CTe", "Valor Teste", "Situação"],
                "E65100")

            ws4 = wb.create_sheet("Só no RPT")
            _escrever_aba(ws4,
                [[d["numero"], d["valor"], d.get("situacao", "")] for d in so_rpt],
                ["Número CTe", "Valor RPT", "Situação"],
                "1565C0")

            wb.save(path)
            messagebox.showinfo("Exportado!", f"Resultado salvo em:\n{path}")
        except Exception as ex:
            messagebox.showerror("Erro ao exportar", str(ex))


if __name__ == "__main__":
    app = App()
    app.mainloop()